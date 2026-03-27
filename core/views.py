"""Views for the Airport Operations Management System.

This module provides views for fiscal assessments, reports, and document management.

Rules compliance: PEP8, type hints, docstrings.
"""

import json
from datetime import datetime, timedelta
from typing import Any, Dict, List, Optional

from django.contrib.auth.mixins import LoginRequiredMixin
from django.core.exceptions import ValidationError
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.db.models import Avg, Count, Q, Sum
from django.http import Http404, HttpRequest, HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.views import View
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from django.views.decorators.cache import cache_page

from .models import (
    Airport,
    AssessmentPeriod,
    AssessmentStatus,
    Document,
    EventLog,
    FiscalAssessment,
    Flight,
    FlightStatus,
    Gate,
    Report,
    ReportFormat,
    ReportType,
)

from .permissions import (
    PermissionMixin,
    UserRole,
    has_minimum_role,
    can_create_assessment,
    can_edit_assessment,
    can_approve_assessment,
    can_create_report,
    can_create_document,
    IsViewerOrAbove,
    IsEditorOrAbove,
    FiscalAssessmentPermissions,
)
from .forms import (
    FiscalAssessmentCreateForm,
    FiscalAssessmentUpdateForm,
    FiscalAssessmentApprovalForm,
    ReportCreateForm,
    DocumentCreateForm,
    FISCAL_ASSESSMENT_CREATE_FIELDS,
    FISCAL_ASSESSMENT_UPDATE_FIELDS,
)



# BASE VIEW CLASSES 

class BaseListView(LoginRequiredMixin, View):
    """Base class for list views with common pagination and filtering.
    
    Usage:
        class FiscalAssessmentListView(BaseListView):
            model = FiscalAssessment
            template_name = 'core/fiscal_assessment_list.html'
            context_object_name = 'assessments'
            filter_fields = ['airport', 'period_type', 'status']
            
            def get_queryset(self):
                return FiscalAssessment.objects.select_related('airport').all()
    """
    
    model = None
    template_name = None
    context_object_name = 'objects'
    paginate_by = 25
    filter_fields = []
    
    def get_queryset(self):
        """Override this method to provide custom queryset."""
        if self.model:
            return self.model.objects.all()
        return super().get_queryset()
    
    def get_filters(self):
        """Extract filter parameters from request GET params."""
        filters = {}
        for field in self.filter_fields:
            value = self.request.GET.get(field)
            if value:
                filters[field] = value
        return filters
    
    def get_context_data(self, queryset):
        """Build common context for list views."""
        # Apply filters
        filters = self.get_filters()
        if filters:
            queryset = queryset.filter(**filters)
        
        # Ordering
        order_by = self.request.GET.get('order_by', '-created_at')
        queryset = queryset.order_by(order_by)
        
        # Pagination
        paginate_by = int(self.request.GET.get('per_page', self.paginate_by))
        paginator = Paginator(queryset, paginate_by)
        page = self.request.GET.get('page', 1)
        
        try:
            page_obj = paginator.page(page)
        except PageNotAnInteger:
            page_obj = paginator.page(1)
        except EmptyPage:
            page_obj = paginator.page(paginator.num_pages)
        
        context = {
            self.context_object_name: page_obj,
            'paginator': paginator,
            'page_obj': page_obj,
            'per_page': paginate_by,
            'order_by': order_by,
        }
        context.update(filters)
        return context
    
    def get(self, request: HttpRequest) -> render:
        """Display the list view."""
        queryset = self.get_queryset()
        context = self.get_context_data(queryset)
        return render(request, self.template_name, context)


class BaseDetailView(LoginRequiredMixin, View):
    """Base class for detail views.
    
    Usage:
        class FiscalAssessmentDetailView(BaseDetailView):
            model = FiscalAssessment
            template_name = 'core/fiscal_assessment_detail.html'
    """
    
    model = None
    template_name = None
    context_object_name = 'object'
    
    def get_object(self):
        """Get the object to display."""
        pk = self.kwargs.get('pk') or self.kwargs.get(f'{self.context_object_name}_id')
        return get_object_or_404(self.model, pk=pk)
    
    def get_context_data(self, obj):
        """Build context for detail view."""
        return {self.context_object_name: obj}
    
    def get(self, request: HttpRequest, **kwargs) -> render:
        """Display the detail view."""
        obj = self.get_object()
        context = self.get_context_data(obj)
        return render(request, self.template_name, context)


class BaseCreateView(PermissionMixin, View):
    """Base class for create views with common patterns.
    
    Usage:
        class FiscalAssessmentCreateView(BaseCreateView):
            model = FiscalAssessment
            template_name = 'core/fiscal_assessment_form.html'
            form_class = FiscalAssessmentForm
            success_url = '/assessments/'
    """
    
    model = None
    template_name = None
    form_class = None
    success_url = None
    required_role = UserRole.EDITOR
    
    def get_form(self):
        """Get form instance."""
        return self.form_class(self.request.POST if self.request.method == 'POST' else None)
    
    def get_context_data(self, form=None, errors=None):
        """Build context for form view."""
        context = {'form': form or self.get_form()}
        if errors:
            context['errors'] = errors
        return context
    
    def get(self, request: HttpRequest) -> render:
        """Display the create form."""
        return render(request, self.template_name, self.get_context_data())
    
    def post(self, request: HttpRequest) -> redirect:
        """Handle form submission."""
        form = self.get_form()
        if form.is_valid():
            obj = form.save()
            return redirect(self.get_success_url(obj))
        return render(request, self.template_name, self.get_context_data(form, form.errors))
    
    def get_success_url(self, obj=None):
        """Get redirect URL after successful creation."""
        if self.success_url:
            return self.success_url
        return f'{self.model._meta.model_name}_detail'


class BaseUpdateView(PermissionMixin, View):
    """Base class for update views.
    
    Usage:
        class FiscalAssessmentUpdateView(BaseUpdateView):
            model = FiscalAssessment
            template_name = 'core/fiscal_assessment_form.html'
            form_class = FiscalAssessmentForm
    """
    
    model = None
    template_name = None
    form_class = None
    required_role = UserRole.EDITOR
    
    def get_object(self):
        """Get the object to update."""
        pk = self.kwargs.get('pk') or self.kwargs.get(f'{self.model._meta.model_name}_id')
        return get_object_or_404(self.model, pk=pk)
    
    def get_form(self, obj):
        """Get form instance."""
        return self.form_class(
            self.request.POST if self.request.method == 'POST' else None,
            instance=obj
        )
    
    def get_context_data(self, form=None, obj=None, errors=None):
        """Build context for form view."""
        context = {
            'form': form or self.get_form(obj),
            self.model._meta.model_name: obj,
        }
        if errors:
            context['errors'] = errors
        return context
    
    def get(self, request: HttpRequest, **kwargs) -> render:
        """Display the edit form."""
        obj = self.get_object()
        return render(request, self.template_name, self.get_context_data(obj=obj))
    
    def post(self, request: HttpRequest, **kwargs) -> redirect:
        """Handle form submission."""
        obj = self.get_object()
        form = self.get_form(obj)
        if form.is_valid():
            form.save()
            return redirect('core:fiscal_assessment_detail', assessment_id=obj.pk)
        return render(request, self.template_name, self.get_context_data(form, obj, form.errors))


class DashboardView(PermissionMixin, View):
    """Main dashboard view showing airport operations overview.

    Displays flight status summary, gate utilization, recent flights,
    fiscal assessment summary, KPIs, alerts, and trend charts.
    """

    template_name = "core/dashboard.html"
    # All authenticated users can view the dashboard (VIEWER role)
    required_role = UserRole.VIEWER

    def get(self, request: HttpRequest) -> render:
        """Display the main dashboard with flight and assessment data."""
        from datetime import datetime, timedelta
        from django.db.models import Sum, Avg
        from .weather_service import weather_service
        from .map_service import map_service
        
        # Get filter parameters
        filter_airport = request.GET.get('airport', '')
        filter_date = request.GET.get('date', '')
        filter_status = request.GET.get('status', '')
        
        # Calculate date range for filters
        today = timezone.now().date()
        if filter_date:
            try:
                filter_date_obj = datetime.strptime(filter_date, '%Y-%m-%d').date()
            except ValueError:
                filter_date_obj = today
        else:
            filter_date_obj = today
        
        # Base flight queryset with filters
        flights_qs = Flight.objects.select_related("gate")
        if filter_airport:
            flights_qs = flights_qs.filter(
                Q(origin=filter_airport) | Q(destination=filter_airport)
            )
        if filter_status:
            flights_qs = flights_qs.filter(status=filter_status)
        
        # Optimized: Get flight statistics by status in a single query
        status_counts = Flight.objects.aggregate(
            **{status.value: Count("id", filter=Q(status=status.value)) for status in FlightStatus}
        )
        flight_status_counts = status_counts
        
        # Get recent flights (last 10)
        recent_flights = flights_qs.order_by("-scheduled_departure")[:10]
        
        # Get upcoming flights (scheduled for today/tomorrow)
        tomorrow = today + timedelta(days=1)
        upcoming_flights = Flight.objects.filter(
            scheduled_departure__date__gte=today,
            scheduled_departure__date__lt=tomorrow
        ).select_related("gate").order_by("scheduled_departure")[:10]
        
        # Get gate utilization
        gates = Gate.objects.annotate(
            flight_count=Count("flights", filter=Q(
                flights__status__in=[
                    FlightStatus.SCHEDULED.value,
                    FlightStatus.BOARDING.value,
                    FlightStatus.DEPARTED.value
                ]
            ))
        ).order_by("gate_id")
        
        # Get recent fiscal assessments
        recent_assessments = FiscalAssessment.objects.select_related(
            "airport"
        ).order_by("-created_at")[:5]
        
        # Get airports for filter
        airports = Airport.objects.filter(is_active=True).order_by("code")
        
        # ============ NEW: KPI Calculations ============
        total_flights = Flight.objects.count()
        total_gates = Gate.objects.count()
        
        # On-time performance (flights not delayed or cancelled)
        on_time_flights = Flight.objects.exclude(
            status__in=['delayed', 'cancelled']
        ).count()
        on_time_rate = round((on_time_flights / total_flights * 100), 1) if total_flights > 0 else 0
        
        # Gate utilization rate
        active_gates = gates.filter(flight_count__gt=0).count()
        gate_utilization_rate = round((active_gates / total_gates * 100), 1) if total_gates > 0 else 0
        
        # Passenger throughput (from assessments)
        total_passengers = FiscalAssessment.objects.aggregate(
            total=Sum('passenger_count')
        )['total'] or 0
        
        # Daily flight trend (today vs yesterday)
        today_flights = Flight.objects.filter(
            scheduled_departure__date=today
        ).count()
        yesterday = today - timedelta(days=1)
        yesterday_flights = Flight.objects.filter(
            scheduled_departure__date=yesterday
        ).count()
        daily_trend = round(((today_flights - yesterday_flights) / yesterday_flights * 100), 1) if yesterday_flights > 0 else 0
        
        # Average delay minutes
        avg_delay = Flight.objects.filter(
            status='delayed'
        ).aggregate(avg=Avg('delay_minutes'))['avg'] or 0
        
        # ============ NEW: 7-Day Flight Trend Data ============
        seven_days_ago = today - timedelta(days=6)
        daily_flights = Flight.objects.filter(
            scheduled_departure__date__gte=seven_days_ago
        ).extra(
            select={'date': 'DATE(scheduled_departure)'}
        ).values('date').annotate(
            count=Count('id')
        ).order_by('date')
        
        # Format for chart: list of [date, count]
        flight_trend_data = []
        date_labels = []
        for i in range(7):
            d = seven_days_ago + timedelta(days=i)
            date_labels.append(d.strftime('%a'))  # Mon, Tue, etc.
            # Find count for this date
            count = 0
            for item in daily_flights:
                if item['date'] == d:
                    count = item['count']
                    break
            flight_trend_data.append(count)
        
        # ============ NEW: Hourly Traffic Data ============
        today_flights_qs = Flight.objects.filter(
            scheduled_departure__date=today
        )
        hourly_data = [0] * 24
        for flight in today_flights_qs:
            hour = flight.scheduled_departure.hour
            hourly_data[hour] += 1
        
        # ============ NEW: Alerts / Action Items ============
        # Flights requiring attention
        delayed_flights_alert = Flight.objects.filter(
            status='delayed'
        ).select_related('gate').order_by('-scheduled_departure')[:5]
        
        cancelled_flights_alert = Flight.objects.filter(
            status='cancelled'
        ).select_related('gate').order_by('-scheduled_departure')[:5]
        
        # Pending assessments awaiting approval
        pending_assessments = FiscalAssessment.objects.filter(
            status='pending'
        ).select_related('airport').order_by('created_at')[:5]
        
        # Gate conflicts (flights with same gate and overlapping times)
        gate_conflicts = []
        for gate in Gate.objects.all():
            gate_flights = Flight.objects.filter(
                gate=gate,
                scheduled_departure__date=today
            ).order_by('scheduled_departure')
            if gate_flights.count() > 1:
                # Check for overlaps
                flights_list = list(gate_flights)
                for i in range(len(flights_list) - 1):
                    if flights_list[i].scheduled_arrival > flights_list[i+1].scheduled_departure:
                        gate_conflicts.append({
                            'gate': gate,
                            'flight1': flights_list[i],
                            'flight2': flights_list[i+1]
                        })
        
        # ============ NEW: Weather Data (Real-time from Open-Meteo) ============
        # Get weather for the primary airport or default
        primary_airport = airports.first() if airports.exists() else None
        if primary_airport:
            weather_data = weather_service.get_weather(
                airport_code=primary_airport.code,
                latitude=float(primary_airport.latitude) if primary_airport.latitude else None,
                longitude=float(primary_airport.longitude) if primary_airport.longitude else None
            )
        else:
            weather_data = weather_service.get_weather('ATL')  # Default to Atlanta
        
        # ============ NEW: Map Data ============
        map_config = map_service.get_map_config()
        active_flights = map_service.get_active_flights(filter_airport if filter_airport else None)
        gates_data = map_service.get_gates_data()
        airports_data = map_service.get_airports_data()
        
        from .permissions import has_minimum_role, UserRole
        
        context = {
            # Existing context
            "flight_status_counts": flight_status_counts,
            "recent_flights": recent_flights,
            "upcoming_flights": upcoming_flights,
            "gates": gates,
            "recent_assessments": recent_assessments,
            "airports": airports,
            "total_flights": total_flights,
            "total_gates": total_gates,
            
            # NEW: KPIs
            "on_time_rate": on_time_rate,
            "gate_utilization_rate": gate_utilization_rate,
            "total_passengers": total_passengers,
            "daily_trend": daily_trend,
            "avg_delay": round(avg_delay, 1),
            "today_flights": today_flights,
            "yesterday_flights": yesterday_flights,
            "active_gates": active_gates,
            
            # NEW: Chart data
            "flight_trend_labels": json.dumps(date_labels),
            "flight_trend_data": json.dumps(flight_trend_data),
            "hourly_data": json.dumps(hourly_data),
            
            # NEW: Alerts
            "delayed_flights_alert": delayed_flights_alert,
            "cancelled_flights_alert": cancelled_flights_alert,
            "pending_assessments": pending_assessments,
            "gate_conflicts": gate_conflicts,
            "alert_count": len(delayed_flights_alert) + len(cancelled_flights_alert) + len(pending_assessments) + len(gate_conflicts),
            
            # NEW: Weather
            "weather_data": weather_data,
            
            # NEW: Map data
            "map_config": json.dumps(map_config),
            "active_flights": json.dumps(active_flights),
            "gates_data": json.dumps(gates_data),
            "airports_data": json.dumps(airports_data),
            
            # NEW: Filters
            "filter_airport": filter_airport,
            "filter_date": filter_date,
            "filter_status": filter_status,
            "today": today,
            
            # Permission flags for template
            "can_create_assessment": can_create_assessment(request.user),
            "can_edit_assessment": can_edit_assessment(request.user),
            "can_approve_assessment": can_approve_assessment(request.user),
            "can_create_report": can_create_report(request.user),
            "can_create_document": can_create_document(request.user),
            "can_access_admin": request.user.is_superuser,
            "user_role": UserRole.EDITOR if can_create_assessment(request.user) else UserRole.VIEWER,
        }
        
        return render(request, self.template_name, context)


class FiscalAssessmentListView(LoginRequiredMixin, View):
    """View for listing all fiscal assessments.
    
    Displays a table of all fiscal assessments with filtering options
    by airport, period type, and status.
    """
    
    template_name = "core/fiscal_assessment_list.html"
    
    def get(self, request: HttpRequest) -> render:
        """Display the list of fiscal assessments."""
        airports = Airport.objects.filter(is_active=True).order_by("code")
        assessments = FiscalAssessment.objects.select_related("airport").all()
        
        # Apply filters
        airport_id = request.GET.get("airport")
        period_type = request.GET.get("period_type")
        status = request.GET.get("status")
        
        if airport_id:
            assessments = assessments.filter(airport_id=airport_id)
        if period_type:
            assessments = assessments.filter(period_type=period_type)
        if status:
            assessments = assessments.filter(status=status)
        
        assessments = assessments.order_by("-start_date")
        
        # Pagination - default 25 items per page
        paginate_by = int(request.GET.get('per_page', 25))
        paginator = Paginator(assessments, paginate_by)
        page = request.GET.get('page', 1)
        
        try:
            assessments_page = paginator.page(page)
        except PageNotAnInteger:
            assessments_page = paginator.page(1)
        except EmptyPage:
            assessments_page = paginator.page(paginator.num_pages)
        
        context = {
            "assessments": assessments_page,
            "airports": airports,
            "selected_airport": airport_id,
            "selected_period_type": period_type,
            "selected_status": status,
            "period_choices": AssessmentPeriod.values,
            "status_choices": AssessmentStatus.values,
            "per_page": paginate_by,
            "paginator": paginator,
        }
        return render(request, self.template_name, context)


class FiscalAssessmentDetailView(LoginRequiredMixin, View):
    """View for displaying fiscal assessment details.
    
    Shows comprehensive details of a single fiscal assessment including
    all financial metrics, revenue breakdown, and expense breakdown.
    """
    
    template_name = "core/fiscal_assessment_detail.html"
    
    def get(self, request: HttpRequest, assessment_id: int) -> render:
        """Display detailed view of a fiscal assessment."""
        assessment = get_object_or_404(
            FiscalAssessment.objects.select_related("airport"),
            id=assessment_id
        )
        
        context = {
            "assessment": assessment,
            "airport": assessment.airport,
        }
        return render(request, self.template_name, context)


class FiscalAssessmentCreateView(PermissionMixin, View):
    """View for creating a new fiscal assessment.
    
    Provides a form for creating new fiscal assessments with options
    for different period types and date ranges.
    """
    
    template_name = "core/fiscal_assessment_form.html"
    required_role = UserRole.EDITOR
    
    def get(self, request: HttpRequest) -> render:
        """Display the create form for fiscal assessment."""
        airports = Airport.objects.filter(is_active=True).order_by("code")
        
        context = {
            "airports": airports,
            "period_choices": AssessmentPeriod.values,
            "status_choices": AssessmentStatus.values,
        }
        return render(request, self.template_name, context)
    
    def post(self, request: HttpRequest) -> redirect:
        """Create a new fiscal assessment with proper validation."""
        form = FiscalAssessmentCreateForm(request.POST)
        
        if not form.is_valid():
            # Return form errors
            airports = Airport.objects.filter(is_active=True).order_by("code")
            context = {
                "airports": airports,
                "period_choices": AssessmentPeriod.values,
                "status_choices": AssessmentStatus.values,
                "errors": form.errors,
            }
            return render(request, self.template_name, context)
        
        try:
            cleaned_data = form.cleaned_data
            
            # Create the assessment with validated data
            assessment = FiscalAssessment.objects.create(
                airport=cleaned_data['airport'],
                period_type=cleaned_data['period_type'],
                start_date=cleaned_data['start_date'],
                end_date=cleaned_data['end_date'],
                status=cleaned_data.get('status', AssessmentStatus.DRAFT.value),
                fuel_revenue=cleaned_data.get('fuel_revenue', 0) or 0,
                parking_revenue=cleaned_data.get('parking_revenue', 0) or 0,
                retail_revenue=cleaned_data.get('retail_revenue', 0) or 0,
                landing_fees=cleaned_data.get('landing_fees', 0) or 0,
                cargo_revenue=cleaned_data.get('cargo_revenue', 0) or 0,
                other_revenue=cleaned_data.get('other_revenue', 0) or 0,
                security_costs=cleaned_data.get('security_costs', 0) or 0,
                maintenance_costs=cleaned_data.get('maintenance_costs', 0) or 0,
                operational_costs=cleaned_data.get('operational_costs', 0) or 0,
                staff_costs=cleaned_data.get('staff_costs', 0) or 0,
                utility_costs=cleaned_data.get('utility_costs', 0) or 0,
                other_expenses=cleaned_data.get('other_expenses', 0) or 0,
                passenger_count=cleaned_data.get('passenger_count', 0) or 0,
                flight_count=cleaned_data.get('flight_count', 0) or 0,
                assessment_notes=cleaned_data.get('assessment_notes', ''),
                assessed_by=cleaned_data.get('assessed_by', ''),
            )
            assessment.calculate_totals().save()
            
            return redirect("fiscal_assessment_detail", assessment_id=assessment.id)
            
        except Exception as e:
            airports = Airport.objects.filter(is_active=True).order_by("code")
            context = {
                "airports": airports,
                "period_choices": AssessmentPeriod.values,
                "status_choices": AssessmentStatus.values,
                "error": str(e),
            }
            return render(request, self.template_name, context)


class FiscalAssessmentUpdateView(PermissionMixin, View):
    """View for updating an existing fiscal assessment."""
    
    template_name = "core/fiscal_assessment_form.html"
    required_role = UserRole.EDITOR
    
    def get(self, request: HttpRequest, assessment_id: int) -> render:
        """Display the edit form for fiscal assessment."""
        assessment = get_object_or_404(FiscalAssessment, id=assessment_id)
        airports = Airport.objects.filter(is_active=True).order_by("code")
        
        context = {
            "assessment": assessment,
            "airports": airports,
            "period_choices": AssessmentPeriod.values,
            "status_choices": AssessmentStatus.values,
        }
        return render(request, self.template_name, context)
    
    def post(self, request: HttpRequest, assessment_id: int) -> redirect:
        """Update an existing fiscal assessment with form validation and field allowlist."""
        assessment = get_object_or_404(FiscalAssessment, id=assessment_id)
        
        # Use form for validation - only allowlisted fields can be updated
        form = FiscalAssessmentUpdateForm(request.POST)
        
        if not form.is_valid():
            airports = Airport.objects.filter(is_active=True).order_by("code")
            context = {
                "assessment": assessment,
                "airports": airports,
                "period_choices": AssessmentPeriod.values,
                "status_choices": AssessmentStatus.values,
                "errors": form.errors,
            }
            return render(request, self.template_name, context)
        
        try:
            cleaned_data = form.cleaned_data
            
            # Only update fields that are in the allowlist and have values
            for field_name in FISCAL_ASSESSMENT_UPDATE_FIELDS:
                if field_name in cleaned_data and cleaned_data[field_name] is not None:
                    # Handle airport specially - it's a ModelChoiceField
                    if field_name == 'airport':
                        setattr(assessment, 'airport', cleaned_data[field_name])
                    else:
                        setattr(assessment, field_name, cleaned_data[field_name])
            
            assessment.calculate_totals().save()
            
            return redirect("fiscal_assessment_detail", assessment_id=assessment.id)
            
        except Exception as e:
            airports = Airport.objects.filter(is_active=True).order_by("code")
            context = {
                "assessment": assessment,
                "airports": airports,
                "period_choices": AssessmentPeriod.values,
                "status_choices": AssessmentStatus.values,
                "error": str(e),
            }
            return render(request, self.template_name, context)


class FiscalAssessmentApproveView(PermissionMixin, View):
    """View for approving a fiscal assessment.
    
    This view includes:
    - Rate limiting to prevent abuse
    - Permission checks to verify user has approval rights
    - Audit logging for all approval actions
    """
    
    # Minimum role required to access this view
    required_role = UserRole.APPROVER
    
    # Rate limiting: minimum seconds between approvals (5 minutes)
    RATE_LIMIT_SECONDS = 300
    # Cache key prefix for rate limiting
    RATE_LIMIT_PREFIX = "approve_rate_limit_"
    
    def _check_rate_limit(self, request: HttpRequest) -> bool:
        """Check if the user has exceeded the rate limit for approvals.
        
        Args:
            request: The HTTP request.
            
        Returns:
            True if within rate limit, False if exceeded.
        """
        from django.core.cache import cache
        
        cache_key = f"{self.RATE_LIMIT_PREFIX}{request.user.id}"
        last_approval = cache.get(cache_key)
        
        if last_approval:
            time_since_last = (timezone.now() - last_approval).total_seconds()
            if time_since_last < self.RATE_LIMIT_SECONDS:
                return False
        
        return True
    
    def _update_rate_limit(self, request: HttpRequest) -> None:
        """Update the rate limit cache for the user."""
        from django.core.cache import cache
        
        cache_key = f"{self.RATE_LIMIT_PREFIX}{request.user.id}"
        cache.set(cache_key, timezone.now(), self.RATE_LIMIT_SECONDS)
    
    def _has_approval_permission(self, request: HttpRequest) -> bool:
        """Check if the user has permission to approve assessments.
        
        Args:
            request: The HTTP request.
            
        Returns:
            True if user has permission, False otherwise.
        """
        user = request.user
        
        # Check if user is a superuser
        if user.is_superuser:
            return True
        
        # Check if user is in the 'approvers' group
        if user.groups.filter(name='approvers').exists():
            return True
        
        # Check if user has specific permission
        if user.has_perm('core.change_fiscalassessment'):
            return True
        
        return False
    
    def _log_approval(self, request: HttpRequest, assessment: FiscalAssessment, 
                      approved: bool, error: str = None) -> None:
        """Log approval actions for audit trail.
        
        Args:
            request: The HTTP request.
            assessment: The fiscal assessment being approved/rejected.
            approved: Whether the assessment was approved.
            error: Optional error message if the operation failed.
        """
        import logging
        
        audit_logger = logging.getLogger('fiscal_assessment_audit')
        
        if error:
            audit_logger.warning(
                f"APPROVAL_FAILED: user={request.user.username}, "
                f"assessment_id={assessment.id}, error={error}"
            )
        else:
            action = "APPROVED" if approved else "REJECTED"
            audit_logger.info(
                f"ASSESSMENT_{action}: user={request.user.username}, "
                f"assessment_id={assessment.id}, airport={assessment.airport.code}, "
                f"period={assessment.period_type}, status={assessment.status}"
            )
    
    def post(self, request: HttpRequest, assessment_id: int) -> redirect:
        """Approve a fiscal assessment with rate limiting and permission checks."""
        # Check rate limit first
        if not self._check_rate_limit(request):
            from django.contrib import messages
            messages.error(
                request, 
                "Please wait before submitting another approval. Try again in a few minutes."
            )
            return redirect("fiscal_assessment_detail", assessment_id=assessment_id)
        
        # Check permission
        if not self._has_approval_permission(request):
            from django.contrib import messages
            from django.http import HttpResponseForbidden
            
            self._log_approval(request, 
                get_object_or_404(FiscalAssessment, id=assessment_id),
                False, 
                "Permission denied"
            )
            messages.error(request, "You do not have permission to approve fiscal assessments.")
            return redirect("fiscal_assessment_detail", assessment_id=assessment_id)
        
        assessment = get_object_or_404(FiscalAssessment, id=assessment_id)
        
        # Validate approval request
        form = FiscalAssessmentApprovalForm(request.POST)
        if not form.is_valid():
            from django.contrib import messages
            messages.error(request, "Invalid approval request.")
            return redirect("fiscal_assessment_detail", assessment_id=assessment_id)
        
        try:
            approved = form.cleaned_data.get('approved', False)
            
            if approved:
                if assessment.status != AssessmentStatus.APPROVED.value:
                    old_status = assessment.status
                    assessment.status = AssessmentStatus.APPROVED.value
                    assessment.approved_by = request.user.username
                    assessment.approved_at = timezone.now()
                    assessment.save()
                    
                    # Log successful approval
                    self._log_approval(request, assessment, True)
                    
                    from django.contrib import messages
                    messages.success(request, f"Assessment approved successfully.")
            else:
                # Handle rejection
                assessment.status = AssessmentStatus.DRAFT.value
                assessment.save()
                
                self._log_approval(request, assessment, False)
                
                from django.contrib import messages
                messages.info(request, "Assessment returned to draft.")
            
            # Update rate limit
            self._update_rate_limit(request)
            
        except Exception as e:
            self._log_approval(request, assessment, False, str(e))
            from django.contrib import messages
            messages.error(request, f"Error processing approval: {str(e)}")
        
        return redirect("fiscal_assessment_detail", assessment_id=assessment.id)


class FiscalAssessmentPrintView(LoginRequiredMixin, View):
    """View for printing fiscal assessment."""
    
    template_name = "core/fiscal_assessment_print.html"
    
    def get(self, request: HttpRequest, assessment_id: int) -> render:
        """Display printable view of a fiscal assessment."""
        assessment = get_object_or_404(
            FiscalAssessment.objects.select_related("airport"),
            id=assessment_id
        )
        
        context = {
            "assessment": assessment,
            "airport": assessment.airport,
        }
        return render(request, self.template_name, context)


class ReportListView(LoginRequiredMixin, View):
    """View for listing all reports.
    
    Displays a table of all generated reports with filtering options
    by airport, report type, and format.
    """
    
    template_name = "core/report_list.html"
    
    def get(self, request: HttpRequest) -> render:
        """Display the list of reports."""
        airports = Airport.objects.filter(is_active=True).order_by("code")
        reports = Report.objects.select_related("airport").all()
        
        # Apply filters
        airport_id = request.GET.get("airport")
        report_type = request.GET.get("report_type")
        report_format = request.GET.get("format")
        
        if airport_id:
            reports = reports.filter(airport_id=airport_id)
        if report_type:
            reports = reports.filter(report_type=report_type)
        if report_format:
            reports = reports.filter(format=report_format)
        
        reports = reports.order_by("-created_at")
        
        # Pagination - default 25 items per page
        paginate_by = int(request.GET.get('per_page', 25))
        paginator = Paginator(reports, paginate_by)
        page = request.GET.get('page', 1)
        
        try:
            reports_page = paginator.page(page)
        except PageNotAnInteger:
            reports_page = paginator.page(1)
        except EmptyPage:
            reports_page = paginator.page(paginator.num_pages)
        
        context = {
            "reports": reports_page,
            "airports": airports,
            "selected_airport": airport_id,
            "selected_report_type": report_type,
            "selected_format": report_format,
            "report_type_choices": ReportType.values,
            "format_choices": ReportFormat.values,
            "per_page": paginate_by,
            "paginator": paginator,
        }
        return render(request, self.template_name, context)


class ReportCreateView(PermissionMixin, View):
    """View for creating a new report.
    
    Provides a form for generating new reports with options
    for different report types and date ranges.
    """
    
    template_name = "core/report_form.html"
    required_role = UserRole.EDITOR
    
    def get(self, request: HttpRequest) -> render:
        """Display the create form for report."""
        airports = Airport.objects.filter(is_active=True).order_by("code")
        form = ReportCreateForm()
        
        context = {
            "form": form,
            "airports": airports,
            "report_type_choices": ReportType.values,
            "format_choices": ReportFormat.values,
        }
        return render(request, self.template_name, context)
    
    def post(self, request: HttpRequest) -> redirect:
        """Create a new report and queue for background generation."""
        form = ReportCreateForm(request.POST)
        if not form.is_valid():
            airports = Airport.objects.filter(is_active=True).order_by("code")
            context = {
                "form": form,
                "airports": airports,
                "report_type_choices": ReportType.values,
                "format_choices": ReportFormat.values,
                "errors": form.errors,
            }
            return render(request, self.template_name, context)
            
        try:
            cleaned_data = form.cleaned_data
            generated_by = request.user.username if request.user.is_authenticated else "system"
            
            # Create the report record
            report = Report.objects.create(
                airport=cleaned_data['airport'],
                report_type=cleaned_data['report_type'],
                title=cleaned_data['title'],
                description=cleaned_data.get('description', ''),
                period_start=cleaned_data['period_start'],
                period_end=cleaned_data['period_end'],
                format=cleaned_data.get('format', ReportFormat.HTML.value),
                generated_by=generated_by,
            )
            
            # Queue for background generation
            from django_q.tasks import async_task
            async_task('core.tasks.generate_report_task', report.id)
            
            from django.contrib import messages
            messages.success(request, f"Report '{report.title}' is being generated in the background.")
            
            return redirect("core:report_list")
            
        except Exception as e:
            airports = Airport.objects.filter(is_active=True).order_by("code")
            return render(request, self.template_name, {
                "form": form,
                "airports": airports,
                "report_type_choices": ReportType.values,
                "format_choices": ReportFormat.values,
                "error": str(e)
            })


class TrendDataAPIView(LoginRequiredMixin, View):
    """API endpoint for dashboard trend data (for Chart.js)."""
    
    def get(self, request: HttpRequest) -> JsonResponse:
        """Get trend data for revenue and passenger counts."""
        # Get last 6 months of data
        six_months_ago = timezone.now().date() - timedelta(days=180)
        
        assessments = FiscalAssessment.objects.filter(
            start_date__gte=six_months_ago
        ).order_by('start_date')
        
        labels = []
        revenue_data = []
        passenger_data = []
        
        for a in assessments:
            labels.append(f"{a.airport.code} ({a.start_date.strftime('%b %Y')})")
            revenue_data.append(float(a.total_revenue))
            passenger_data.append(a.passenger_count)
            
        return JsonResponse({
            "labels": labels,
            "revenue": revenue_data,
            "passengers": passenger_data,
        })
    
    def _generate_report_content(self, report: Report) -> Dict[str, Any]:
        """Generate content for the report based on its type."""
        from core.models import Flight, Passenger, FiscalAssessment
        
        content = {
            "generated_at": timezone.now().isoformat(),
            "period": {
                "start": report.period_start.isoformat(),
                "end": report.period_end.isoformat(),
            },
        }
        
        if report.report_type == ReportType.FISCAL_SUMMARY.value:
            # Get fiscal assessments for the period
            assessments = FiscalAssessment.objects.filter(
                airport=report.airport,
                start_date__gte=report.period_start,
                end_date__lte=report.period_end,
            )
            
            total_revenue = sum(a.total_revenue for a in assessments)
            total_expenses = sum(a.total_expenses for a in assessments)
            net_profit = sum(a.net_profit for a in assessments)
            total_passengers = sum(a.passenger_count for a in assessments)
            total_flights = sum(a.flight_count for a in assessments)
            
            content["summary"] = {
                "total_revenue": float(total_revenue),
                "total_expenses": float(total_expenses),
                "net_profit": float(net_profit),
                "total_passengers": total_passengers,
                "total_flights": total_flights,
                "assessment_count": assessments.count(),
            }
            content["assessments"] = [
                {
                    "period": f"{a.start_date} to {a.end_date}",
                    "type": a.period_type,
                    "revenue": float(a.total_revenue),
                    "expenses": float(a.total_expenses),
                    "profit": float(a.net_profit),
                }
                for a in assessments
            ]
            
        elif report.report_type == ReportType.OPERATIONAL.value:
            # Get flights for the period
            flights = Flight.objects.filter(
                gate__isnull=False,
                scheduled_departure__gte=report.period_start,
                scheduled_departure__lte=report.period_end,
            )
            
            total_flights = flights.count()
            delayed_flights = flights.filter(status="delayed").count()
            cancelled_flights = flights.filter(status="cancelled").count()
            avg_delay = flights.aggregate(Avg("delay_minutes"))["delay_minutes__avg"] or 0
            
            content["summary"] = {
                "total_flights": total_flights,
                "delayed_flights": delayed_flights,
                "cancelled_flights": cancelled_flights,
                "on_time_flights": total_flights - delayed_flights - cancelled_flights,
                "average_delay_minutes": round(avg_delay, 2),
            }
            
        elif report.report_type == ReportType.PASSENGER.value:
            # Get passengers for the period
            passengers = Passenger.objects.filter(
                flight__scheduled_departure__gte=report.period_start,
                flight__scheduled_departure__lte=report.period_end,
            )
            
            total_passengers = passengers.count()
            checked_in = passengers.filter(status="checked_in").count()
            boarded = passengers.filter(status="boarded").count()
            arrived = passengers.filter(status="arrived").count()
            no_show = passengers.filter(status="no_show").count()
            
            content["summary"] = {
                "total_passengers": total_passengers,
                "checked_in": checked_in,
                "boarded": boarded,
                "arrived": arrived,
                "no_show": no_show,
            }
            
        elif report.report_type == ReportType.FINANCIAL.value:
            # Similar to fiscal but with different formatting
            assessments = FiscalAssessment.objects.filter(
                airport=report.airport,
                start_date__gte=report.period_start,
                end_date__lte=report.period_end,
            )
            
            revenue_breakdown = {
                "fuel_revenue": sum(float(a.fuel_revenue) for a in assessments),
                "parking_revenue": sum(float(a.parking_revenue) for a in assessments),
                "retail_revenue": sum(float(a.retail_revenue) for a in assessments),
                "landing_fees": sum(float(a.landing_fees) for a in assessments),
                "cargo_revenue": sum(float(a.cargo_revenue) for a in assessments),
                "other_revenue": sum(float(a.other_revenue) for a in assessments),
            }
            
            expense_breakdown = {
                "security_costs": sum(float(a.security_costs) for a in assessments),
                "maintenance_costs": sum(float(a.maintenance_costs) for a in assessments),
                "operational_costs": sum(float(a.operational_costs) for a in assessments),
                "staff_costs": sum(float(a.staff_costs) for a in assessments),
                "utility_costs": sum(float(a.utility_costs) for a in assessments),
                "other_expenses": sum(float(a.other_expenses) for a in assessments),
            }
            
            content["revenue_breakdown"] = revenue_breakdown
            content["expense_breakdown"] = expense_breakdown
            content["net_profit"] = sum(revenue_breakdown.values()) - sum(expense_breakdown.values())
        
        return content


class ReportDetailView(LoginRequiredMixin, View):
    """View for displaying report details."""
    
    template_name = "core/report_detail.html"
    
    def get(self, request: HttpRequest, report_id: int) -> render:
        """Display detailed view of a report."""
        report = get_object_or_404(
            Report.objects.select_related("airport"),
            id=report_id
        )
        
        context = {
            "report": report,
            "airport": report.airport,
        }
        return render(request, self.template_name, context)


class ReportExportView(LoginRequiredMixin, View):
    """View for exporting reports to different formats."""

    def get(self, request: HttpRequest, report_id: int) -> HttpRequest:
        """Export report in specified format."""
        report = get_object_or_404(Report, id=report_id)
        export_format = request.GET.get("format", "json")

        if export_format == "json":
            return self._export_json(report)
        elif export_format == "csv":
            return self._export_csv(report)
        elif export_format == "xlsx":
            return self._export_xlsx(report)
        else:
            return JsonResponse({"error": "Invalid format"}, status=400)

    def _export_json(self, report: Report) -> JsonResponse:
        """Export report as JSON."""
        data = {
            "report_id": str(report.report_id),
            "title": report.title,
            "airport": report.airport.code,
            "report_type": report.report_type,
            "period_start": str(report.period_start),
            "period_end": str(report.period_end),
            "generated_at": report.generated_at.isoformat() if report.generated_at else None,
            "content": report.content,
        }
        response = JsonResponse(data, json_dumps_params={"indent": 2})
        response["Content-Disposition"] = f'attachment; filename="{report.title.replace(" ", "_")}.json"'
        return response

    def _export_csv(self, report: Report) -> HttpRequest:
        """Export report as CSV."""
        import csv
        import io

        output = io.StringIO()
        writer = csv.writer(output)

        # Write header
        writer.writerow(["Field", "Value"])
        writer.writerow(["Report ID", str(report.report_id)])
        writer.writerow(["Title", report.title])
        writer.writerow(["Airport", report.airport.code])
        writer.writerow(["Type", report.report_type])
        writer.writerow(["Period Start", str(report.period_start)])
        writer.writerow(["Period End", str(report.period_end)])
        writer.writerow(["Generated By", report.generated_by])
        writer.writerow(["Generated At", str(report.generated_at) if report.generated_at else ""])

        # Write content if available
        if report.content:
            writer.writerow([])
            writer.writerow(["Content"])
            if "summary" in report.content:
                summary = report.content["summary"]
                if isinstance(summary, dict):
                    for key, value in summary.items():
                        writer.writerow([key, value])
                elif isinstance(summary, str):
                    writer.writerow(["Summary", summary])

        response = HttpResponse(output.getvalue(), content_type="text/csv")
        response["Content-Disposition"] = f'attachment; filename="{report.title.replace(" ", "_")}.csv"'
        return response

    def _export_xlsx(self, report: Report) -> HttpResponse:
        """Export report as Excel XLSX."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        import io

        wb = Workbook()
        ws = wb.active
        ws.title = "Report"

        # Styles
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="0d6efd", end_color="0d6efd", fill_type="solid")
        title_font = Font(bold=True, size=14)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Header
        ws.merge_cells('A1:B1')
        ws['A1'] = f"Report: {report.title}"
        ws['A1'].font = title_font

        # Basic info
        info = [
            ("Report ID", str(report.report_id)),
            ("Title", report.title),
            ("Airport", report.airport.code),
            ("Type", report.report_type),
            ("Period Start", str(report.period_start)),
            ("Period End", str(report.period_end)),
            ("Generated By", report.generated_by or ""),
            ("Generated At", str(report.generated_at) if report.generated_at else ""),
        ]

        for row_idx, (label, value) in enumerate(info, start=3):
            ws.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row_idx, column=2, value=value)

        # Content section
        if report.content:
            ws.append([])
            ws.append(["Content Summary"])
            ws.cell(row=ws.max_row, column=1).font = header_font
            ws.cell(row=ws.max_row, column=1).fill = header_fill

            if "summary" in report.content:
                summary = report.content["summary"]
                if isinstance(summary, dict):
                    # Header row
                    header_row = ws.max_row + 1
                    ws.cell(row=header_row, column=1, value="Metric")
                    ws.cell(row=header_row, column=2, value="Value")
                    for col in range(1, 3):
                        cell = ws.cell(row=header_row, column=col)
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = Alignment(horizontal='center')

                    # Data rows
                    for key, value in summary.items():
                        ws.append([key.replace('_', ' ').title(), value])

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save to buffer
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(output.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response["Content-Disposition"] = f'attachment; filename="{report.title.replace(" ", "_")}.xlsx"'
        return response


class DocumentListView(LoginRequiredMixin, View):
    """View for listing all documents."""
    
    template_name = "core/document_list.html"
    
    def get(self, request: HttpRequest) -> render:
        """Display the list of documents."""
        airports = Airport.objects.filter(is_active=True).order_by("code")
        documents = Document.objects.select_related("airport").all()
        
        # Apply filters
        airport_id = request.GET.get("airport")
        document_type = request.GET.get("document_type")
        
        if airport_id:
            documents = documents.filter(airport_id=airport_id)
        if document_type:
            documents = documents.filter(document_type=document_type)
        
        documents = documents.order_by("-created_at")
        
        # Pagination - default 25 items per page
        paginate_by = int(request.GET.get('per_page', 25))
        paginator = Paginator(documents, paginate_by)
        page = request.GET.get('page', 1)
        
        try:
            documents_page = paginator.page(page)
        except PageNotAnInteger:
            documents_page = paginator.page(1)
        except EmptyPage:
            documents_page = paginator.page(paginator.num_pages)
        
        context = {
            "documents": documents_page,
            "airports": airports,
            "selected_airport": airport_id,
            "selected_document_type": document_type,
            "document_type_choices": Document.DocumentType.values,
            "per_page": paginate_by,
            "paginator": paginator,
        }
        return render(request, self.template_name, context)


class DocumentCreateView(PermissionMixin, View):
    """View for creating a new document."""
    
    template_name = "core/document_form.html"
    required_role = UserRole.EDITOR
    
    def get(self, request: HttpRequest) -> render:
        """Display the create form for document."""
        airports = Airport.objects.filter(is_active=True).order_by("code")
        form = DocumentCreateForm()
        
        context = {
            "form": form,
            "airports": airports,
            "document_type_choices": Document.DocumentType.values,
        }
        return render(request, self.template_name, context)
    
    def post(self, request: HttpRequest) -> redirect:
        """Create a new document with validation."""
        form = DocumentCreateForm(request.POST)
        if not form.is_valid():
            airports = Airport.objects.filter(is_active=True).order_by("code")
            context = {
                "form": form,
                "airports": airports,
                "document_type_choices": Document.DocumentType.values,
                "errors": form.errors,
            }
            return render(request, self.template_name, context)
            
        try:
            cleaned_data = form.cleaned_data
            created_by = request.user.username if request.user.is_authenticated else "system"
            
            document = Document.objects.create(
                name=cleaned_data['name'],
                document_type=cleaned_data['document_type'],
                airport=cleaned_data['airport'],
                content=cleaned_data['content'],
                is_template=cleaned_data.get('is_template', False),
                created_by=created_by,
            )
            
            return redirect("core:document_detail", document_id=document.id)
            
        except Exception as e:
            airports = Airport.objects.filter(is_active=True).order_by("code")
            return render(request, self.template_name, {
                "form": form,
                "airports": airports,
                "document_type_choices": Document.DocumentType.values,
                "error": str(e)
            })


class DocumentDetailView(LoginRequiredMixin, View):
    """View for displaying document details."""
    
    template_name = "core/document_detail.html"
    
    def get(self, request: HttpRequest, document_id: int) -> render:
        """Display detailed view of a document."""
        document = get_object_or_404(
            Document.objects.select_related("airport"),
            id=document_id
        )
        
        context = {
            "document": document,
        }
        return render(request, self.template_name, context)


class DocumentExportView(LoginRequiredMixin, View):
    """View for exporting documents."""
    
    def get(self, request: HttpRequest, document_id: int) -> JsonResponse:
        """Export document as JSON."""
        document = get_object_or_404(Document, id=document_id)
        
        data = {
            "document_id": str(document.document_id),
            "name": document.name,
            "document_type": document.document_type,
            "airport": document.airport.code if document.airport else None,
            "is_template": document.is_template,
            "created_by": document.created_by,
            "created_at": document.created_at.isoformat(),
            "content": document.content,
        }
        
        response = JsonResponse(data, json_dumps_params={"indent": 2})
        response["Content-Disposition"] = f'attachment; filename="{document.name.replace(" ", "_")}.json"'
        return response


# API Views for AJAX operations

class FiscalAssessmentAPIView(LoginRequiredMixin, View):
    """API endpoint for fiscal assessment operations.
    
    CSRF protection is ENABLED (removed csrf_exempt) because this view 
    uses session-based authentication.
    """
    
    def get(self, request: HttpRequest, assessment_id: Optional[int] = None) -> JsonResponse:
        """Get fiscal assessment data."""
        if assessment_id:
            assessment = get_object_or_404(FiscalAssessment, id=assessment_id)
            return JsonResponse({
                "id": assessment.id,
                "airport": assessment.airport.code,
                "period_type": assessment.period_type,
                "start_date": str(assessment.start_date),
                "end_date": str(assessment.end_date),
                "status": assessment.status,
                "total_revenue": float(assessment.total_revenue),
                "total_expenses": float(assessment.total_expenses),
                "net_profit": float(assessment.net_profit),
                "passenger_count": assessment.passenger_count,
                "flight_count": assessment.flight_count,
            })
        else:
            # List all assessments
            assessments = FiscalAssessment.objects.select_related("airport").all()
            data = [
                {
                    "id": a.id,
                    "airport": a.airport.code,
                    "period_type": a.period_type,
                    "start_date": str(a.start_date),
                    "end_date": str(a.end_date),
                    "status": a.status,
                    "total_revenue": float(a.total_revenue),
                    "net_profit": float(a.net_profit),
                }
                for a in assessments
            ]
            return JsonResponse({"assessments": data})
    
    def delete(self, request: HttpRequest, assessment_id: int) -> JsonResponse:
        """Delete a fiscal assessment.
        
        SECURITY: This operation requires admin (superuser) privileges.
        Role bypass is NOT possible - the check is enforced at multiple levels.
        """
        # CRITICAL SECURITY: Check if user is admin (superuser)
        # This cannot be bypassed - the check uses is_superuser flag directly
        if not request.user.is_superuser:
            import logging
            logger = logging.getLogger(__name__)
            logger.warning(
                f"SECURITY: User {request.user.username} attempted to DELETE "
                f"FiscalAssessment {assessment_id} but lacks admin privileges. "
                f"Request DENIED."
            )
            return JsonResponse(
                {"error": "You do not have permission to delete fiscal assessments. "
                          "Only administrators can perform this action."},
                status=403
            )
        
        assessment = get_object_or_404(FiscalAssessment, id=assessment_id)
        assessment.delete()
        return JsonResponse({"success": True, "message": "Assessment deleted"})


class ReportAPIView(LoginRequiredMixin, View):
    """API endpoint for report operations.
    
    CSRF protection is ENABLED (removed csrf_exempt) because this view 
    uses session-based authentication.
    """
    
    def get(self, request: HttpRequest, report_id: Optional[int] = None) -> JsonResponse:
        """Get report data."""
        if report_id:
            report = get_object_or_404(Report, id=report_id)
            return JsonResponse({
                "id": report.id,
                "title": report.title,
                "airport": report.airport.code,
                "report_type": report.report_type,
                "period_start": str(report.period_start),
                "period_end": str(report.period_end),
                "format": report.format,
                "is_generated": report.is_generated,
                "content": report.content,
            })
        else:
            reports = Report.objects.select_related("airport").all()
            data = [
                {
                    "id": r.id,
                    "title": r.title,
                    "airport": r.airport.code,
                    "report_type": r.report_type,
                    "is_generated": r.is_generated,
                }
                for r in reports
            ]
            return JsonResponse({"reports": data})


class DocumentAPIView(LoginRequiredMixin, View):
    """API endpoint for document operations.
    
    CSRF protection is ENABLED (removed csrf_exempt) because this view 
    uses session-based authentication.
    """
    
    def get(self, request: HttpRequest, document_id: Optional[int] = None) -> JsonResponse:
        """Get document data."""
        if document_id:
            document = get_object_or_404(Document, id=document_id)
            return JsonResponse({
                "id": document.id,
                "name": document.name,
                "document_type": document.document_type,
                "airport": document.airport.code if document.airport else None,
                "is_template": document.is_template,
                "content": document.content,
            })
        else:
            documents = Document.objects.select_related("airport").all()
            data = [
                {
                    "id": d.id,
                    "name": d.name,
                    "document_type": d.document_type,
                    "airport": d.airport.code if d.airport else None,
                    "is_template": d.is_template,
                }
                for d in documents
            ]
            return JsonResponse({"documents": data})


class DashboardSummaryView(LoginRequiredMixin, View):
    """API endpoint for dashboard summary data."""
    
    @method_decorator(cache_page(300))
    def get(self, request: HttpRequest) -> JsonResponse:
        """Get summary data for the dashboard."""
        airports = Airport.objects.filter(is_active=True)
        
        # Get recent assessments
        recent_assessments = FiscalAssessment.objects.select_related("airport").order_by("-created_at")[:5]
        
        # Optimized: Get summary stats via database aggregation
        financial_stats = FiscalAssessment.objects.aggregate(
            total_revenue=Sum('total_revenue'),
            total_expenses=Sum('total_expenses'),
            net_profit=Sum('net_profit')
        )
        
        total_revenue = float(financial_stats['total_revenue'] or 0)
        total_expenses = float(financial_stats['total_expenses'] or 0)
        net_profit = float(financial_stats['net_profit'] or 0)
        
        # Get report counts
        report_stats = Report.objects.aggregate(
            generated=Count("id", filter=Q(is_generated=True)),
            pending=Count("id", filter=Q(is_generated=False))
        )
        
        # Get document counts
        document_stats = Document.objects.aggregate(
            total=Count("id"),
            templates=Count("id", filter=Q(is_template=True))
        )
        
        return JsonResponse({
            "airports_count": airports.count(),
            "assessments_count": FiscalAssessment.objects.count(),
            "reports_generated": report_stats['generated'],
            "reports_pending": report_stats['pending'],
            "total_documents": document_stats['total'],
            "templates": document_stats['templates'],
            "financial_summary": {
                "total_revenue": total_revenue,
                "total_expenses": total_expenses,
                "net_profit": net_profit,
            },
            "recent_assessments": [
                {
                    "id": a.id,
                    "airport": a.airport.code,
                    "period_type": a.period_type,
                    "net_profit": float(a.net_profit),
                    "status": a.status,
                }
                for a in recent_assessments
            ],
        })


class EventLogListView(LoginRequiredMixin, View):
    """View for listing all activity/event logs.
    
    Displays a table of all activities carried out in the system,
    with filtering options by event type, action, severity, user, and date range.
    """
    
    template_name = "core/activity_log_list.html"
    
    def get(self, request: HttpRequest) -> render:
        """Display the list of activity logs with filtering options."""
        logs = EventLog.objects.select_related('user', 'flight').all()
        
        # Apply filters
        event_type = request.GET.get("event_type")
        action = request.GET.get("action")
        severity = request.GET.get("severity")
        user_id = request.GET.get("user")
        date_from = request.GET.get("date_from")
        date_to = request.GET.get("date_to")
        search = request.GET.get("search")
        
        if event_type:
            logs = logs.filter(event_type=event_type)
        if action:
            logs = logs.filter(action=action)
        if severity:
            logs = logs.filter(severity=severity)
        if user_id:
            logs = logs.filter(user_id=user_id)
        if date_from:
            logs = logs.filter(timestamp__date__gte=date_from)
        if date_to:
            logs = logs.filter(timestamp__date__lte=date_to)
        if search:
            # Improved search: search across multiple fields with fuzzy matching
            from django.db.models import Q
            logs = logs.filter(
                Q(description__icontains=search) |
                Q(event_type__icontains=search) |
                Q(action__icontains=search) |
                Q(user__username__icontains=search) |
                Q(user__email__icontains=search) |
                Q(flight__flight_number__icontains=search) |
                Q(ip_address__icontains=search)
            )
        
        # Get unique event types and actions for filter dropdowns
        event_types = EventLog.objects.values_list('event_type', flat=True).distinct()
        actions = EventLog.objects.values_list('action', flat=True).distinct()
        severities = [severity.value for severity in EventLog.EventSeverity]
        
        # Order by most recent
        logs = logs.order_by("-timestamp")
        
        # Pagination - default 50 items per page
        paginate_by = int(request.GET.get('per_page', 50))
        paginator = Paginator(logs, paginate_by)
        page = request.GET.get('page', 1)
        
        try:
            logs_page = paginator.page(page)
        except PageNotAnInteger:
            logs_page = paginator.page(1)
        except EmptyPage:
            logs_page = paginator.page(paginator.num_pages)
        
        # Get users for filter dropdown
        from django.contrib.auth.models import User
        users = User.objects.filter(
            id__in=EventLog.objects.values_list('user_id', flat=True).distinct()
        ).order_by('username')
        
        context = {
            "logs": logs_page,
            "event_types": sorted(event_types),
            "actions": sorted(actions),
            "severities": severities,
            "users": users,
            "selected_event_type": event_type,
            "selected_action": action,
            "selected_severity": severity,
            "selected_user": user_id,
            "selected_date_from": date_from,
            "selected_date_to": date_to,
            "search_query": search,
            "per_page": paginate_by,
            "paginator": paginator,
            "total_logs": logs.count(),
        }
        return render(request, self.template_name, context)


class AnalyticsDashboardView(PermissionMixin, View):
    """Analytics dashboard view with interactive charts.

    Provides comprehensive analytics including revenue trends,
    flight statistics, passenger trends, and airport comparisons.
    """

    required_role = UserRole.VIEWER
    template_name = "core/analytics_dashboard.html"

    def get(self, request: HttpRequest) -> HttpResponse:
        """Render analytics dashboard."""
        from .models import Airport, Gate, Flight, Passenger, FiscalAssessment, AssessmentStatus
        from django.db.models import Count, Sum, Avg, Q

        context = {
            "page_title": "Analytics Dashboard",
            "total_airports": Airport.objects.filter(is_active=True).count(),
            "total_gates": Gate.objects.count(),
            "total_flights": Flight.objects.count(),
            "total_passengers": Passenger.objects.count(),
            "approved_assessments": FiscalAssessment.objects.filter(
                status=AssessmentStatus.APPROVED.value
            ).count(),
        }
        return render(request, self.template_name, context)


class ReportScheduleListView(PermissionMixin, View):
    """List view for scheduled reports."""

    required_role = UserRole.EDITOR
    template_name = "core/report_schedule_list.html"

    def get(self, request: HttpRequest) -> HttpResponse:
        """Render scheduled reports list."""
        from .models import ReportSchedule
        from .forms import ReportScheduleForm

        schedules = ReportSchedule.objects.select_related('airport', 'created_by').all()

        context = {
            "page_title": "Scheduled Reports",
            "schedules": schedules,
            "form": ReportScheduleForm(),
        }
        return render(request, self.template_name, context)


class ReportScheduleCreateView(PermissionMixin, View):
    """Create view for report schedules."""

    required_role = UserRole.EDITOR
    template_name = "core/report_schedule_form.html"

    def get(self, request: HttpRequest) -> HttpResponse:
        """Render create form."""
        from .forms import ReportScheduleForm

        context = {
            "page_title": "New Report Schedule",
            "form": ReportScheduleForm(),
            "action": "create",
        }
        return render(request, self.template_name, context)

    def post(self, request: HttpRequest) -> HttpResponse:
        """Handle schedule creation."""
        from .forms import ReportScheduleForm
        from .models import ReportSchedule
        from django.urls import reverse
        from django.contrib import messages

        form = ReportScheduleForm(request.POST)
        if form.is_valid():
            schedule = form.save(commit=False)
            schedule.created_by = request.user
            schedule.save()

            # Queue initial check
            from django_q.tasks import async_task
            async_task('core.tasks.check_scheduled_reports', group='schedules')

            messages.success(request, f"Report schedule '{schedule.name}' created successfully.")
            return redirect(reverse('core:report_schedule_list'))

        context = {
            "page_title": "New Report Schedule",
            "form": form,
            "action": "create",
        }
        return render(request, self.template_name, context, status=400)


class ReportScheduleUpdateView(PermissionMixin, View):
    """Update view for report schedules."""

    required_role = UserRole.EDITOR
    template_name = "core/report_schedule_form.html"

    def get(self, request: HttpRequest, schedule_id: int) -> HttpResponse:
        """Render edit form."""
        from .models import ReportSchedule
        from .forms import ReportScheduleForm

        schedule = get_object_or_404(ReportSchedule, id=schedule_id)

        context = {
            "page_title": "Edit Report Schedule",
            "form": ReportScheduleForm(instance=schedule),
            "action": "update",
            "schedule": schedule,
        }
        return render(request, self.template_name, context)

    def post(self, request: HttpRequest, schedule_id: int) -> HttpResponse:
        """Handle schedule update."""
        from .models import ReportSchedule
        from .forms import ReportScheduleForm
        from django.urls import reverse
        from django.contrib import messages

        schedule = get_object_or_404(ReportSchedule, id=schedule_id)
        form = ReportScheduleForm(request.POST, instance=schedule)

        if form.is_valid():
            form.save()
            messages.success(request, f"Report schedule '{schedule.name}' updated successfully.")
            return redirect(reverse('core:report_schedule_list'))

        context = {
            "page_title": "Edit Report Schedule",
            "form": form,
            "action": "update",
            "schedule": schedule,
        }
        return render(request, self.template_name, context, status=400)


class ReportScheduleDeleteView(PermissionMixin, View):
    """Delete view for report schedules."""

    required_role = UserRole.EDITOR

    def post(self, request: HttpRequest, schedule_id: int) -> HttpResponse:
        """Handle schedule deletion."""
        from .models import ReportSchedule
        from django.urls import reverse
        from django.contrib import messages

        schedule = get_object_or_404(ReportSchedule, id=schedule_id)
        schedule_name = schedule.name
        schedule.delete()

        messages.success(request, f"Report schedule '{schedule_name}' deleted successfully.")
        return redirect(reverse('core:report_schedule_list'))


class ReportScheduleRunNowView(PermissionMixin, View):
    """Manually trigger a scheduled report."""

    required_role = UserRole.EDITOR

    def post(self, request: HttpRequest, schedule_id: int) -> HttpResponse:
        """Trigger immediate report generation."""
        from .models import ReportSchedule
        from django.urls import reverse
        from django.contrib import messages
        from django_q.tasks import async_task

        schedule = get_object_or_404(ReportSchedule, id=schedule_id)

        async_task('core.tasks.generate_scheduled_report', schedule.id, group='schedules')

        messages.success(request, f"Report '{schedule.name}' is being generated.")
        return redirect(reverse('core:report_schedule_list'))


class AirportComparisonView(PermissionMixin, View):
    """Multi-airport comparison view with side-by-side metrics."""

    required_role = UserRole.VIEWER
    template_name = "core/airport_comparison.html"

    def get(self, request: HttpRequest) -> HttpResponse:
        """Render airport comparison page."""
        from .models import Airport, FiscalAssessment, AssessmentStatus, Flight, Gate, Passenger
        from django.db.models import Count, Sum, Avg, Q
        from datetime import timedelta, date

        # Get selected airports from query params
        airport_ids = request.GET.getlist('airports')
        comparison_period = request.GET.get('period', '6months')
        start_date_str = request.GET.get('start_date', '')
        end_date_str = request.GET.get('end_date', '')

        # Calculate date cutoff based on period
        now = timezone.now()
        cutoff = None
        end_date = None
        
        if comparison_period == 'custom' and start_date_str:
            # Custom date range
            try:
                cutoff = timezone.make_aware(date.fromisoformat(start_date_str))
                if end_date_str:
                    end_date = timezone.make_aware(date.fromisoformat(end_date_str))
                else:
                    end_date = now
            except (ValueError, TypeError):
                # Fallback to 6 months if invalid dates
                cutoff = now - timedelta(days=180)
                end_date = now
        elif comparison_period == '7days':
            cutoff = now - timedelta(days=7)
        elif comparison_period == '30days':
            cutoff = now - timedelta(days=30)
        elif comparison_period == '90days':
            cutoff = now - timedelta(days=90)
        elif comparison_period == '6months':
            cutoff = now - timedelta(days=180)
        elif comparison_period == '1year':
            cutoff = now - timedelta(days=365)
        else:
            # Default to 6 months
            cutoff = now - timedelta(days=180)

        # Get all active airports
        all_airports = Airport.objects.filter(is_active=True)

        # Filter selected airports
        if airport_ids:
            airports = all_airports.filter(id__in=airport_ids)
        else:
            airports = all_airports[:4]  # Default to first 4 airports

        # Build comparison data
        comparison_data = []
        for airport in airports:
            # Filter assessments by date range
            assessment_filters = Q(
                airport=airport,
                status=AssessmentStatus.APPROVED.value,
                start_date__gte=cutoff
            )
            if end_date:
                assessment_filters &= Q(start_date__lte=end_date)
            
            assessments = FiscalAssessment.objects.filter(assessment_filters)
            
            # Flight model uses airport codes (origin/destination), not FK
            flight_filters = Q(
                (Q(origin=airport.code) | Q(destination=airport.code)),
                scheduled_departure__gte=cutoff
            )
            if end_date:
                flight_filters &= Q(scheduled_departure__lte=end_date)
            
            flights = Flight.objects.filter(flight_filters)

            total_revenue = assessments.aggregate(Sum('total_revenue'))['total_revenue__sum'] or 0
            total_expenses = assessments.aggregate(Sum('total_expenses'))['total_expenses__sum'] or 0
            net_profit = assessments.aggregate(Sum('net_profit'))['net_profit__sum'] or 0
            total_passengers = assessments.aggregate(Sum('passenger_count'))['passenger_count__sum'] or 0
            total_flights = flights.count()
            delayed_flights = flights.filter(status='delayed').count()
            on_time_rate = ((total_flights - delayed_flights) / total_flights * 100) if total_flights > 0 else 0
            profit_margin = (net_profit / total_revenue * 100) if total_revenue > 0 else 0

            comparison_data.append({
                'airport': {
                    'id': airport.id,
                    'code': airport.code,
                    'name': airport.name,
                },
                'total_revenue': float(total_revenue),
                'total_expenses': float(total_expenses),
                'net_profit': float(net_profit),
                'total_passengers': total_passengers,
                'total_flights': total_flights,
                'on_time_rate': round(on_time_rate, 1),
                'assessment_count': assessments.count(),
                'profit_margin': round(profit_margin, 1),
            })

        context = {
            "page_title": "Airport Comparison",
            "all_airports": all_airports,
            "selected_airport_ids": [int(a) for a in airport_ids] if airport_ids else [a.id for a in airports],
            "comparison_period": comparison_period,
            "comparison_data": comparison_data,
            "start_date": start_date_str,
            "end_date": end_date_str,
        }
        return render(request, self.template_name, context)


class FlightStatusPortalView(View):
    """Public flight status portal for passengers.
    
    This view does not require authentication and provides
    real-time flight status information for passengers.
    """

    template_name = "core/public/flight_status.html"

    def get(self, request: HttpRequest) -> HttpResponse:
        """Render flight status portal."""
        from .models import Flight, FlightStatus, Airport, Baggage
        from django.db.models import Q

        search_query = request.GET.get('q', '')
        status_filter = request.GET.get('status', '')
        airport_filter = request.GET.get('airport', '')

        # Start with all flights today
        today = timezone.now().date()
        flights = Flight.objects.filter(
            Q(scheduled_departure__date=today) | Q(scheduled_arrival__date=today)
        ).select_related('gate').order_by('scheduled_departure')

        # Apply filters
        if search_query:
            flights = flights.filter(
                Q(flight_number__icontains=search_query) |
                Q(origin__icontains=search_query) |
                Q(destination__icontains=search_query) |
                Q(airline__icontains=search_query)
            )

        if status_filter:
            flights = flights.filter(status=status_filter)

        if airport_filter:
            flights = flights.filter(
                Q(origin=airport_filter) | Q(destination=airport_filter)
            )

        # Get status choices for filter
        status_choices = [(status.value, status.name.replace('_', ' ').title()) for status in FlightStatus]
        
        # Get airports for filter
        airports = Airport.objects.filter(is_active=True)

        # Demo hints (random examples to help first-time users)
        demo_flight_hint = None
        demo_baggage_hint = None
        if not search_query:
            import random

            demo_flights_qs = Flight.objects.filter(
                Q(scheduled_departure__date=today) | Q(scheduled_arrival__date=today)
            ).values_list('flight_number', flat=True)
            flight_count = demo_flights_qs.count()
            if flight_count:
                demo_flight_hint = demo_flights_qs[random.randrange(flight_count)]

            demo_baggage_qs = Baggage.objects.values_list('tag_number', flat=True)
            baggage_count = demo_baggage_qs.count()
            if baggage_count:
                demo_baggage_hint = demo_baggage_qs[random.randrange(baggage_count)]

        context = {
            "page_title": "Flight Status",
            "flights": flights[:50],  # Limit to 50 for performance
            "status_choices": status_choices,
            "airports": airports,
            "selected_status": status_filter,
            "selected_airport": airport_filter,
            "search_query": search_query,
            "demo_flight_hint": demo_flight_hint,
            "demo_baggage_hint": demo_baggage_hint,
            "is_public": True,
        }
        return render(request, self.template_name, context)


class BaggageTrackingView(View):
    """Public baggage tracking portal for passengers."""

    template_name = "core/public/baggage_tracking.html"

    def get(self, request: HttpRequest) -> HttpResponse:
        """Render baggage tracking page."""
        from .models import Baggage, Flight
        from django.db.models import Q

        tag_number = request.GET.get('tag', '')
        baggage = None
        error = None

        if tag_number:
            try:
                baggage = Baggage.objects.select_related('passenger', 'flight').get(
                    tag_number=tag_number.upper()
                )
            except Baggage.DoesNotExist:
                error = f"Baggage with tag number '{tag_number}' not found."

        # Demo hints (random examples to help first-time users)
        demo_baggage_hint = None
        demo_flight_hint = None
        if not tag_number:
            import random

            demo_baggage_qs = Baggage.objects.values_list('tag_number', flat=True)
            baggage_count = demo_baggage_qs.count()
            if baggage_count:
                demo_baggage_hint = demo_baggage_qs[random.randrange(baggage_count)]

            today = timezone.now().date()
            demo_flights_qs = Flight.objects.filter(
                Q(scheduled_departure__date=today) | Q(scheduled_arrival__date=today)
            ).values_list('flight_number', flat=True)
            flight_count = demo_flights_qs.count()
            if flight_count:
                demo_flight_hint = demo_flights_qs[random.randrange(flight_count)]

        context = {
            "page_title": "Baggage Tracking",
            "baggage": baggage,
            "error": error,
            "tag_number": tag_number,
            "demo_baggage_hint": demo_baggage_hint,
            "demo_flight_hint": demo_flight_hint,
            "is_public": True,
        }
        return render(request, self.template_name, context)


class DataImportWizardView(LoginRequiredMixin, View):
    """Guided CSV import wizard with column mapping."""

    template_name = "core/import_wizard.html"

    def get(self, request: HttpRequest, model_name: str) -> HttpResponse:
        """Render import wizard."""
        from .models import Flight, Passenger, Staff, Gate, Airport
        
        models = {
            'flight': Flight,
            'passenger': Passenger,
            'staff': Staff,
            'gate': Gate,
            'airport': Airport,
        }
        
        if model_name not in models:
            raise Http404("Model not found")
        
        model = models[model_name]
        fields = []
        for field in model._meta.fields:
            if field.editable and field.name not in ['id', 'created_at', 'updated_at']:
                fields.append({
                    'name': field.name,
                    'label': field.verbose_name.title() if hasattr(field, 'verbose_name') else field.name.replace('_', ' ').title()
                })
        
        context = {
            "page_title": f"Import {model_name.title()}s",
            "model_name": model_name,
            "fields": json.dumps(fields),
            "back_url": request.META.get('HTTP_REFERER', '/'),
        }
        return render(request, self.template_name, context)

    def post(self, request: HttpRequest, model_name: str) -> JsonResponse:
        """Process CSV import."""
        import csv
        import io
        from .models import Flight, Passenger, Staff, Gate, Airport
        
        models = {
            'flight': Flight,
            'passenger': Passenger,
            'staff': Staff,
            'gate': Gate,
            'airport': Airport,
        }
        
        if model_name not in models:
            return JsonResponse({'success': False, 'error': 'Invalid model'})
        
        model = models[model_name]
        mapping = json.loads(request.POST.get('mapping', '{}'))
        
        if not mapping:
            return JsonResponse({'success': False, 'error': 'No column mapping provided'})
        
        try:
            csv_file = request.FILES.get('file')
            if not csv_file:
                return JsonResponse({'success': False, 'error': 'No file uploaded'})
            
            decoded_file = csv_file.read().decode('utf-8')
            reader = csv.DictReader(decoded_file.splitlines())
            
            imported = 0
            errors = []
            
            for row_num, row in enumerate(reader, start=2):
                try:
                    data = {}
                    for csv_col, model_field in mapping.items():
                        if csv_col in row and row[csv_col]:
                            data[model_field] = row[csv_col]
                    
                    if data:
                        model.objects.create(**data)
                        imported += 1
                except Exception as e:
                    errors.append(f"Row {row_num}: {str(e)}")
            
            return JsonResponse({
                'success': True,
                'imported': imported,
                'errors': errors[:10]  # Limit errors shown
            })
            
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})


def health_check(request):
    """Lightweight health check endpoint for monitoring/self-ping."""
    return JsonResponse({"status": "ok"})
