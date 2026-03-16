"""Admin configuration for the Airport Operations Management System.

ADMIN ACCESS CONTROL - Least Privilege Principle:
The Django admin panel is restricted to superusers only.
Regular users (including editors and approvers) cannot access admin.
This follows the principle of least privilege.
"""

import csv
import io
import logging
from datetime import datetime
from django.contrib import admin
from django.http import HttpRequest, HttpResponse, HttpResponseRedirect, HttpResponseForbidden
from django.shortcuts import render
from django.urls import path


from .models import (
    Airport,
    Document,
    EventLog,
    FiscalAssessment,
    Flight,
    Gate,
    Passenger,
    Report,
    Staff,
    StaffAssignment,
    ReportSchedule,
    Shift,
    StaffShiftAssignment,
    Baggage,
    WeatherCondition,
    WeatherAlert,
    FuelInventory,
    FuelDispensing,
    MaintenanceSchedule,
    MaintenanceLog,
    CustomField,
    CustomFieldValue,
)


def validate_csv_file(csv_file, max_size_mb=2):
    """Validate CSV file for import using whitelist approach.
    
    Args:
        csv_file: The uploaded file to validate.
        max_size_mb: Maximum file size in megabytes.
    
    Returns:
        tuple: (is_valid, error_message)
    """
    # Get whitelist from settings - only allow specific extensions
    from django.conf import settings
    allowed_extensions = getattr(settings, 'ALLOWED_FILE_EXTENSIONS', ['.csv'])
    
    # Check file extension against whitelist
    file_ext = '.' + csv_file.name.split('.')[-1].lower() if '.' in csv_file.name else ''
    if file_ext not in allowed_extensions:
        return False, f"Invalid file type. Only {', '.join(allowed_extensions)} files are allowed."
    
    # Check file size (max by parameter or settings default)
    max_size = getattr(settings, 'MAX_UPLOAD_SIZE', max_size_mb * 1024 * 1024)
    if csv_file.size > max_size:
        return False, f"File too large. Maximum size is {max_size // (1024*1024)}MB."
    
    return True, None


def sanitize_import_error(row_num: int, exception: Exception, logger: logging.Logger = None) -> str:
    """Sanitize error messages from CSV imports to prevent information disclosure.
    
    This function creates safe error messages for users while logging the actual
    errors for debugging purposes.
    
    Args:
        row_num: The row number where the error occurred.
        exception: The original exception.
        logger: Optional logger for recording actual errors.
    
    Returns:
        A sanitized error message safe to display to users.
    """
    # Log the actual error for debugging 
    if logger:
        logger.warning(f"CSV import error at row {row_num}: {type(exception).__name__}: {str(exception)}")
    
    # Return sanitized generic message
    return f"Row {row_num}: Invalid data format"


# Get logger for CSV import operations
csv_import_logger = logging.getLogger('csv_import')


# CUSTOM ADMIN SITE - Restrict Access to Superusers Only
class RestrictedAdminSite(admin.AdminSite):
    """Custom admin site that restricts access to superusers only."""
    
    def has_permission(self, request):
        """Check if user has permission to access admin."""
        return request.user.is_active and request.user.is_superuser
    
    def admin_view(self, view, cacheable=False):
        """Wrap the admin view to check for superuser status."""
        def wrapper(request, *args, **kwargs):
            if not request.user.is_superuser:
                return HttpResponseForbidden(
                    "Access denied. Only superusers can access the admin panel."
                )
            return view(request, *args, **kwargs)
        return wrapper

# Instantiate the custom admin site
admin_site = RestrictedAdminSite(name='restricted_admin')


# MIXINS - Reusable functionality for admin classes
class CSVImportMixin:
    """Mixin that provides common CSV import functionality.
    
    Subclasses must implement:
    - get_model(): Returns the model class
    - process_row(row, row_num): Process a single CSV row, returns (object, error_msg)
    - get_model_name(): Returns lowercase model name for templates
    """
    
    def get_urls(self):
        urls = super().get_urls()
        return [
            path('import-csv/', self.import_csv, name='import_csv'),
        ] + urls
    
    def _check_superuser(self, request: HttpRequest) -> bool:
        """Check if user is superuser, return True if authorized."""
        if not request.user.is_superuser:
            self.message_user(request, "Only superusers can import CSV data.", level='error')
            return False
        return True
    
    def _process_csv_file(self, csv_file, process_row_func):
        """Process CSV file and return created_count and errors."""
        data_set = csv_file.read().decode('UTF-8')
        io_string = io.StringIO(data_set)
        reader = csv.reader(io_string)
        
        # Skip header
        next(reader)
        
        created_count = 0
        errors = []
        for row_num, row in enumerate(reader, start=2):
            try:
                obj, error_msg = process_row_func(row, row_num)
                if obj:
                    created_count += 1
                if error_msg:
                    errors.append(error_msg)
            except Exception as e:
                errors.append(sanitize_import_error(row_num, e, csv_import_logger))
        
        return created_count, errors
    
    def import_csv(self, request: HttpRequest):
        """Handle CSV import request."""
        # RBAC: Only superusers can import CSV
        if not self._check_superuser(request):
            return HttpResponseRedirect('../')
        
        if request.method == 'POST':
            csv_file = request.FILES.get('csv_file')
            
            # Validate file
            is_valid, error_message = validate_csv_file(csv_file)
            if not is_valid:
                self.message_user(request, error_message, level='error')
                return HttpResponseRedirect('../')
            
            # Process CSV using subclass implementation
            created_count, errors = self._process_csv_file(csv_file, self.process_row)
            
            model_name = self.get_model_name()
            if errors:
                self.message_user(request, f"Imported {created_count} {model_name}. Errors: {'; '.join(errors[:5])}", level='warning')
            else:
                self.message_user(request, f'Imported {created_count} {model_name} successfully!')
            return HttpResponseRedirect('../')
        
        return render(request, 'admin/csv_form.html', {
            'title': 'Import CSV',
            'app_label': 'core',
            'model_name': self.get_model_name()
        })
    
    def process_row(self, row, row_num):
        """Process a single CSV row. Must be implemented by subclass."""
        raise NotImplementedError("Subclasses must implement process_row()")
    
    def get_model_name(self):
        """Return lowercase model name. Must be implemented by subclass."""
        raise NotImplementedError("Subclasses must implement get_model_name()")


class CSVExportMixin:
    """Mixin that provides common CSV export functionality.
    
    Subclasses must implement:
    - get_export_fields(): Returns list of field names for CSV header
    - get_export_row(obj): Returns list of field values for a single object
    """
    
    def export_to_csv(self, request, queryset):
        """Export selected objects to CSV."""
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="{}.csv"'.format(
            self.model._meta.model_name
        )
        
        writer = csv.writer(response)
        writer.writerow(self.get_export_fields())
        
        for obj in queryset:
            writer.writerow(self.get_export_row(obj))
        
        return response
    
    export_to_csv.short_description = "Export selected to CSV"
    
    def get_export_fields(self):
        """Return list of field names for CSV header. Must be implemented by subclass."""
        raise NotImplementedError("Subclasses must implement get_export_fields()")
    
    def get_export_row(self, obj):
        """Return list of field values for a single object. Must be implemented by subclass."""
        raise NotImplementedError("Subclasses must implement get_export_row()")


# ADMIN CLASSES
class PassengerInline(admin.TabularInline):
    """Inline admin for passengers."""
    model = Passenger
    extra = 0
    readonly_fields = ["passenger_id", "created_at", "updated_at"]


@admin.register(Airport, site=admin_site)
class AirportAdmin(CSVImportMixin, admin.ModelAdmin):
    """Admin for Airport model."""
    list_display = ["code", "name", "city", "is_active"]
    search_fields = ["code", "name", "city"]
    list_filter = ["is_active"]
    
    def process_row(self, row, row_num):
        if len(row) >= 3:
            Airport.objects.get_or_create(
                code=row[0].strip().upper(),
                defaults={
                    'name': row[1].strip(),
                    'city': row[2].strip(),
                    'timezone': row[3].strip() if len(row) > 3 else 'UTC',
                }
            )
            return True, None
        return False, f"Row {row_num}: Invalid data format"
    
    def get_model_name(self):
        return "airports"


@admin.register(Gate, site=admin_site)
class GateAdmin(CSVImportMixin, admin.ModelAdmin):
    """Admin for Gate model."""
    list_display = ["gate_id", "terminal", "capacity", "status", "updated_at"]
    list_filter = ["status", "terminal", "capacity"]
    search_fields = ["gate_id", "terminal"]
    readonly_fields = ["created_at", "updated_at"]
    ordering = ["gate_id"]
    
    def process_row(self, row, row_num):
        if len(row) >= 4:
            Gate.objects.get_or_create(
                gate_id=row[0].strip(),
                defaults={
                    'terminal': row[1].strip(),
                    'capacity': row[2].strip(),
                    'status': row[3].strip()
                }
            )
            return True, None
        return False, f"Row {row_num}: Invalid data format"
    
    def get_model_name(self):
        return "gates"


@admin.register(Flight, site=admin_site)
class FlightAdmin(CSVImportMixin, admin.ModelAdmin):
    """Admin for Flight model."""
    list_display = [
        "flight_number", "origin", "destination",
        "scheduled_departure", "gate", "status", "delay_minutes"
    ]
    list_filter = ["status", "origin", "destination"]
    search_fields = ["flight_number", "origin", "destination"]
    readonly_fields = ["created_at", "updated_at"]
    inlines = [PassengerInline]
    ordering = ["scheduled_departure"]
    
    def process_row(self, row, row_num):
        if len(row) >= 8:
            # Get gate if provided
            gate = None
            if len(row) > 6 and row[6].strip():
                gate = Gate.objects.filter(gate_id=row[6].strip()).first()
            
            # Parse datetime
            scheduled_departure = row[4].strip()
            scheduled_arrival = row[5].strip()
            
            Flight.objects.get_or_create(
                flight_number=row[0].strip().upper(),
                defaults={
                    'airline': row[1].strip() if len(row) > 1 else '',
                    'origin': row[2].strip().upper(),
                    'destination': row[3].strip().upper(),
                    'scheduled_departure': scheduled_departure,
                    'scheduled_arrival': scheduled_arrival,
                    'gate': gate,
                    'status': row[7].strip().lower() if len(row) > 7 else 'scheduled',
                    'delay_minutes': int(row[8].strip()) if len(row) > 8 and row[8].strip().isdigit() else 0
                }
            )
            return True, None
        return False, f"Row {row_num}: Invalid data format"
    
    def get_model_name(self):
        return "flights"


@admin.register(Passenger, site=admin_site)
class PassengerAdmin(admin.ModelAdmin):
    """Admin for Passenger model."""
    list_display = ["last_name", "first_name", "passport_number", "flight", "seat_number", "status"]
    list_filter = ["status", "flight"]
    search_fields = ["first_name", "last_name", "passport_number", "seat_number"]
    readonly_fields = ["passenger_id", "created_at", "updated_at"]
    ordering = ["last_name", "first_name"]


@admin.register(Staff, site=admin_site)
class StaffAdmin(CSVImportMixin, admin.ModelAdmin):
    """Admin for Staff model."""
    list_display = ["last_name", "first_name", "employee_number", "role", "is_available"]
    list_filter = ["role", "is_available"]
    search_fields = ["first_name", "last_name", "employee_number"]
    readonly_fields = ["staff_id", "created_at", "updated_at"]
    ordering = ["last_name", "first_name"]
    
    def process_row(self, row, row_num):
        if len(row) >= 5:
            Staff.objects.get_or_create(
                employee_number=row[2].strip(),
                defaults={
                    'first_name': row[0].strip(),
                    'last_name': row[1].strip(),
                    'role': row[3].strip(),
                    'certification': row[4].strip() if len(row) > 4 else '',
                    'is_available': row[5].strip().lower() == 'true' if len(row) > 5 else True,
                }
            )
            return True, None
        return False, f"Row {row_num}: Invalid data format"
    
    def get_model_name(self):
        return "staff"


@admin.register(StaffAssignment, site=admin_site)
class StaffAssignmentAdmin(admin.ModelAdmin):
    """Admin for StaffAssignment model."""
    list_display = ["staff", "flight", "assignment_type", "assigned_at"]
    list_filter = ["assignment_type", "assigned_at"]
    search_fields = ["staff__first_name", "staff__last_name", "flight__flight_number"]
    readonly_fields = ["assigned_at"]
    raw_id_fields = ["staff", "flight"]
    ordering = ["-assigned_at"]


@admin.register(EventLog, site=admin_site)
class EventLogAdmin(admin.ModelAdmin):
    """Admin for EventLog model."""
    list_display = ["timestamp", "event_type", "action", "user", "description", "severity"]
    list_filter = ["severity", "event_type", "action", "timestamp"]
    search_fields = ["description", "event_type", "user__username"]
    readonly_fields = ["event_id", "created_at", "timestamp"]
    ordering = ["-timestamp"]
    date_hierarchy = "timestamp"
    raw_id_fields = ["user", "flight"]
    list_per_page = 50
    
    actions = ["export_to_csv", "export_to_xlsx", "export_audit_log"]

    def export_to_csv(self, request, queryset):
        """Export selected audit logs to CSV."""
        import csv
        import io
        from django.http import HttpResponse

        output = io.StringIO()
        writer = csv.writer(output)
        
        # Write header
        writer.writerow([
            'Event ID', 'Timestamp', 'Event Type', 'Action', 'Severity',
            'User', 'Flight', 'Description', 'IP Address'
        ])
        
        # Write data
        for log in queryset:
            writer.writerow([
                str(log.event_id),
                log.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
                log.event_type,
                log.action,
                log.severity,
                log.user.username if log.user else '',
                log.flight.flight_number if log.flight else '',
                log.description,
                log.ip_address or ''
            ])
        
        response = HttpResponse(output.getvalue(), content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="audit_logs.csv"'
        self.message_user(request, f"Exported {queryset.count()} audit log entries to CSV.")
        return response
    export_to_csv.short_description = "Export selected to CSV"

    def export_to_xlsx(self, request, queryset):
        """Export selected audit logs to Excel."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        import io
        from django.http import HttpResponse

        wb = Workbook()
        ws = wb.active
        ws.title = "Audit Logs"
        
        # Styles
        header_font = Font(bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="4a5568", end_color="4a5568", fill_type="solid")
        
        # Headers
        headers = [
            'Event ID', 'Timestamp', 'Event Type', 'Action', 'Severity',
            'User', 'Flight', 'Description', 'IP Address'
        ]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # Data rows
        for row_idx, log in enumerate(queryset, start=2):
            ws.cell(row=row_idx, column=1, value=str(log.event_id))
            ws.cell(row=row_idx, column=2, value=log.timestamp.strftime('%Y-%m-%d %H:%M:%S'))
            ws.cell(row=row_idx, column=3, value=log.event_type)
            ws.cell(row=row_idx, column=4, value=log.action)
            ws.cell(row=row_idx, column=5, value=log.severity)
            ws.cell(row=row_idx, column=6, value=log.user.username if log.user else '')
            ws.cell(row=row_idx, column=7, value=log.flight.flight_number if log.flight else '')
            ws.cell(row=row_idx, column=8, value=log.description[:200] if log.description else '')
            ws.cell(row=row_idx, column=9, value=log.ip_address or '')
        
        # Auto-adjust column widths
        column_widths = [38, 20, 20, 15, 12, 15, 15, 60, 15]
        for col, width in enumerate(column_widths, start=1):
            ws.column_dimensions[chr(64 + col)].width = width
        
        # Save to buffer
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(output.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="audit_logs.xlsx"'
        self.message_user(request, f"Exported {queryset.count()} audit log entries to Excel.")
        return response
    export_to_xlsx.short_description = "Export selected to Excel"

    def export_audit_log(self, request, queryset):
        """Export full audit log for compliance."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        import io
        from django.http import HttpResponse
        from django.utils import timezone

        wb = Workbook()
        
        # Summary sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"
        
        # Styles
        title_font = Font(bold=True, size=16)
        header_font = Font(bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="2d3748", end_color="2d3748", fill_type="solid")
        
        # Title
        ws_summary.merge_cells('A1:B1')
        ws_summary['A1'] = 'Audit Log Export Report'
        ws_summary['A1'].font = title_font
        
        # Summary info
        now = timezone.now()
        summary_data = [
            ('Export Date', now.strftime('%Y-%m-%d %H:%M:%S')),
            ('Total Entries', queryset.count()),
            ('Date Range', f"{queryset.order_by('timestamp').first().timestamp if queryset.exists() else 'N/A'}"),
            ('Exported By', request.user.username),
        ]
        
        for row_idx, (label, value) in enumerate(summary_data, start=3):
            ws_summary.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
            ws_summary.cell(row=row_idx, column=2, value=value)
        
        # Detail sheet
        ws_detail = wb.create_sheet(title="Audit Logs")
        
        headers = [
            'Timestamp', 'Event ID', 'Type', 'Action', 'Severity',
            'User', 'Flight', 'Description', 'IP Address'
        ]
        for col, header in enumerate(headers, start=1):
            cell = ws_detail.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # Data
        for row_idx, log in enumerate(queryset.order_by('-timestamp'), start=2):
            ws_detail.cell(row=row_idx, column=1, value=log.timestamp.strftime('%Y-%m-%d %H:%M:%S'))
            ws_detail.cell(row=row_idx, column=2, value=str(log.event_id))
            ws_detail.cell(row=row_idx, column=3, value=log.event_type)
            ws_detail.cell(row=row_idx, column=4, value=log.action)
            ws_detail.cell(row=row_idx, column=5, value=log.severity)
            ws_detail.cell(row=row_idx, column=6, value=log.user.username if log.user else 'System')
            ws_detail.cell(row=row_idx, column=7, value=log.flight.flight_number if log.flight else '')
            ws_detail.cell(row=row_idx, column=8, value=log.description[:250] if log.description else '')
            ws_detail.cell(row=row_idx, column=9, value=log.ip_address or '')
        
        # Save
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"audit_log_export_{now.strftime('%Y%m%d_%H%M%S')}.xlsx"
        response = HttpResponse(output.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        self.message_user(request, f"Exported {queryset.count()} audit log entries for compliance.")
        return response
    export_audit_log.short_description = "Export full audit log (compliance)"


@admin.register(FiscalAssessment, site=admin_site)
class FiscalAssessmentAdmin(CSVImportMixin, CSVExportMixin, admin.ModelAdmin):
    """Admin for FiscalAssessment model."""
    list_display = ["airport", "period_type", "start_date", "end_date", "status", "total_revenue", "total_expenses", "net_profit"]
    list_filter = ["status", "period_type", "airport"]
    search_fields = ["airport__code", "airport__name"]
    readonly_fields = ["created_at", "updated_at"]
    ordering = ["-start_date"]
    date_hierarchy = "start_date"
    fieldsets = (
        ("Airport & Period", {
            "fields": ("airport", "period_type", "start_date", "end_date", "status")
        }),
        ("Financial Metrics", {
            "fields": ("total_revenue", "total_expenses", "net_profit")
        }),
        ("Operational Metrics", {
            "fields": ("passenger_count", "flight_count")
        }),
        ("Revenue Breakdown", {
            "fields": ("fuel_revenue", "parking_revenue", "retail_revenue", "landing_fees", "cargo_revenue", "other_revenue")
        }),
        ("Expense Breakdown", {
            "fields": ("security_costs", "maintenance_costs", "operational_costs", "staff_costs", "utility_costs", "other_expenses")
        }),
        ("Approval", {
            "fields": ("assessment_notes", "assessed_by", "approved_by", "approved_at")
        }),
    )
    actions = ["export_to_csv", "approve_selected"]
    
    def process_row(self, row, row_num):
        if len(row) >= 10:
            from core.models import AssessmentStatus
            
            # Get airport
            airport = None
            if row[0].strip():
                airport = Airport.objects.filter(code=row[0].strip().upper()).first()
            
            if not airport:
                return False, f"Row {row_num}: Airport not found"
            
            # Parse dates
            start_date = datetime.strptime(row[2].strip(), '%Y-%m-%d').date() if row[2].strip() else None
            end_date = datetime.strptime(row[3].strip(), '%Y-%m-%d').date() if row[3].strip() else None
            
            # Parse numeric fields
            def parse_decimal(val, default=0):
                try:
                    return float(val.strip()) if val.strip() else default
                except:
                    return default
            
            def parse_int(val, default=0):
                try:
                    return int(val.strip()) if val.strip() else default
                except:
                    return default
            
            FiscalAssessment.objects.create(
                airport=airport,
                period_type=row[1].strip().lower(),
                start_date=start_date,
                end_date=end_date,
                status=row[4].strip().lower() if len(row) > 4 else 'draft',
                total_revenue=parse_decimal(row[5]) if len(row) > 5 else 0,
                total_expenses=parse_decimal(row[6]) if len(row) > 6 else 0,
                net_profit=parse_decimal(row[7]) if len(row) > 7 else 0,
                passenger_count=parse_int(row[8]) if len(row) > 8 else 0,
                flight_count=parse_int(row[9]) if len(row) > 9 else 0,
                assessed_by=self.request.user.username,
            )
            return True, None
        return False, f"Row {row_num}: Invalid data format"
    
    def get_model_name(self):
        return "fiscal assessments"
    
    def _check_superuser(self, request: HttpRequest) -> bool:
        """Check if user is superuser, return True if authorized."""
        # Store request for use in process_row
        self.request = request
        if not request.user.is_superuser:
            self.message_user(request, "Only superusers can import CSV data.", level='error')
            return False
        return True
    
    def get_urls(self):
        urls = super().get_urls()
        return [
            path('import-csv/', self.import_csv, name='import_csv'),
        ] + urls
    
    def get_export_fields(self):
        return [
            'Airport', 'Period Type', 'Start Date', 'End Date', 'Status',
            'Total Revenue', 'Total Expenses', 'Net Profit',
            'Passenger Count', 'Flight Count'
        ]
    
    def get_export_row(self, obj):
        return [
            obj.airport.code,
            obj.period_type,
            obj.start_date,
            obj.end_date,
            obj.status,
            obj.total_revenue,
            obj.total_expenses,
            obj.net_profit,
            obj.passenger_count,
            obj.flight_count,
        ]
    
    def export_to_csv(self, request, queryset):
        """Export selected assessments to CSV."""
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="fiscal_assessments.csv"'
        
        writer = csv.writer(response)
        writer.writerow(self.get_export_fields())
        
        for obj in queryset:
            writer.writerow(self.get_export_row(obj))
        
        return response
    
    def approve_selected(self, request, queryset):
        """Approve selected assessments."""
        from django.utils import timezone
        from core.models import AssessmentStatus
        
        count = 0
        for obj in queryset:
            if obj.status != AssessmentStatus.APPROVED.value:
                obj.status = AssessmentStatus.APPROVED.value
                obj.approved_by = request.user.username
                obj.approved_at = timezone.now()
                obj.save()
                count += 1
        
        self.message_user(request, f"{count} assessments approved.")
    
    approve_selected.short_description = "Approve selected assessments"


@admin.register(Report, site=admin_site)
class ReportAdmin(CSVImportMixin, CSVExportMixin, admin.ModelAdmin):
    """Admin for Report model."""
    list_display = ["title", "airport", "report_type", "period_start", "period_end", "format", "is_generated", "generated_at"]
    list_filter = ["report_type", "format", "is_generated", "airport"]
    search_fields = ["title", "airport__code", "airport__name"]
    readonly_fields = ["report_id", "created_at", "updated_at"]
    ordering = ["-created_at"]
    date_hierarchy = "created_at"
    actions = ["export_to_csv", "regenerate_reports"]
    
    def process_row(self, row, row_num):
        if len(row) >= 6:
            # Get airport
            airport = None
            if row[1].strip():
                airport = Airport.objects.filter(code=row[1].strip().upper()).first()
            
            if not airport:
                return False, f"Row {row_num}: Airport not found"
            
            # Parse dates
            period_start = datetime.strptime(row[3].strip(), '%Y-%m-%d').date() if row[3].strip() else None
            period_end = datetime.strptime(row[4].strip(), '%Y-%m-%d').date() if row[4].strip() else None
            
            Report.objects.create(
                title=row[0].strip(),
                airport=airport,
                report_type=row[2].strip().lower(),
                period_start=period_start,
                period_end=period_end,
                format=row[5].strip().lower() if len(row) > 5 else 'html',
                description=row[6].strip() if len(row) > 6 else '',
                generated_by=self.request.user.username,
            )
            return True, None
        return False, f"Row {row_num}: Invalid data format"
    
    def get_model_name(self):
        return "reports"
    
    def _check_superuser(self, request: HttpRequest) -> bool:
        """Check if user is superuser, return True if authorized."""
        # Store request for use in process_row
        self.request = request
        if not request.user.is_superuser:
            self.message_user(request, "Only superusers can import CSV data.", level='error')
            return False
        return True
    
    def get_urls(self):
        urls = super().get_urls()
        return [
            path('import-csv/', self.import_csv, name='import_csv'),
        ] + urls
    
    def get_export_fields(self):
        return [
            'Title', 'Airport', 'Type', 'Period Start', 'Period End',
            'Format', 'Generated', 'Generated At'
        ]
    
    def get_export_row(self, obj):
        return [
            obj.title,
            obj.airport.code,
            obj.report_type,
            obj.period_start,
            obj.period_end,
            obj.format,
            obj.is_generated,
            obj.generated_at,
        ]
    
    def export_to_csv(self, request, queryset):
        """Export selected reports to CSV."""
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="reports.csv"'
        
        writer = csv.writer(response)
        writer.writerow(self.get_export_fields())
        
        for obj in queryset:
            writer.writerow(self.get_export_row(obj))
        
        return response
    
    def regenerate_reports(self, request, queryset):
        """Regenerate selected reports."""
        from core.views import ReportCreateView
        
        count = 0
        for obj in queryset:
            # Re-generate the report content
            view = ReportCreateView()
            new_content = view._generate_report_content(obj)
            obj.content = new_content
            obj.is_generated = True
            obj.save()
            count += 1
        
        self.message_user(request, f"{count} reports regenerated.")
    
    regenerate_reports.short_description = "Regenerate selected reports"


@admin.register(Document, site=admin_site)
class DocumentAdmin(CSVImportMixin, CSVExportMixin, admin.ModelAdmin):
    """Admin for Document model."""
    list_display = ["name", "document_type", "airport", "is_template", "created_by", "created_at"]
    list_filter = ["document_type", "is_template", "airport"]
    search_fields = ["name", "airport__code", "airport__name"]
    readonly_fields = ["document_id", "created_at", "updated_at"]
    ordering = ["-created_at"]
    actions = ["export_to_csv", "mark_as_template", "mark_as_document"]
    
    def process_row(self, row, row_num):
        if len(row) >= 4:
            # Get airport if provided
            airport = None
            if len(row) > 2 and row[2].strip():
                airport = Airport.objects.filter(code=row[2].strip().upper()).first()
            
            Document.objects.create(
                name=row[0].strip(),
                document_type=row[1].strip().lower(),
                airport=airport,
                content=row[3].strip() if len(row) > 3 else '',
                created_by=self.request.user.username,
            )
            return True, None
        return False, f"Row {row_num}: Invalid data format"
    
    def get_model_name(self):
        return "documents"
    
    def _check_superuser(self, request: HttpRequest) -> bool:
        """Check if user is superuser, return True if authorized."""
        # Store request for use in process_row
        self.request = request
        if not request.user.is_superuser:
            self.message_user(request, "Only superusers can import CSV data.", level='error')
            return False
        return True
    
    def get_urls(self):
        urls = super().get_urls()
        return [
            path('import-csv/', self.import_csv, name='import_csv'),
        ] + urls
    
    def get_export_fields(self):
        return [
            'Name', 'Type', 'Airport', 'Template', 'Created By', 'Created At'
        ]
    
    def get_export_row(self, obj):
        return [
            obj.name,
            obj.document_type,
            obj.airport.code if obj.airport else '',
            obj.is_template,
            obj.created_by,
            obj.created_at,
        ]
    
    def export_to_csv(self, request, queryset):
        """Export selected documents to CSV."""
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="documents.csv"'
        
        writer = csv.writer(response)
        writer.writerow(self.get_export_fields())
        
        for obj in queryset:
            writer.writerow(self.get_export_row(obj))
        
        return response
    
    def mark_as_template(self, request, queryset):
        """Mark selected documents as templates."""
        count = queryset.update(is_template=True)
        self.message_user(request, f"{count} documents marked as templates.")
    
    mark_as_template.short_description = "Mark as template"

    def mark_as_document(self, request, queryset):
        """Mark selected documents as regular documents."""
        count = queryset.update(is_template=False)
        self.message_user(request, f"{count} documents marked as regular documents.")

    mark_as_document.short_description = "Mark as document"


@admin.register(ReportSchedule)
class ReportScheduleAdmin(admin.ModelAdmin):
    """Admin for report schedules."""

    list_display = ['name', 'report_type', 'airport', 'frequency', 'next_run', 'last_run', 'is_active']
    list_filter = ['frequency', 'is_active', 'report_type']
    search_fields = ['name', 'recipients']
    readonly_fields = ['last_run', 'next_run', 'created_by', 'created_at', 'updated_at']

    fieldsets = (
        ('Basic Information', {
            'fields': ('name', 'report_type', 'airport', 'format')
        }),
        ('Schedule', {
            'fields': ('frequency', 'day_of_week', 'day_of_month', 'hour')
        }),
        ('Recipients', {
            'fields': ('recipients',)
        }),
        ('Status', {
            'fields': ('is_active', 'last_run', 'next_run')
        }),
    )

    def save_model(self, request, obj, form, change):
        """Set created_by on save."""
        if not change:
            obj.created_by = request.user
        super().save_model(request, obj, form, change)


@admin.register(Shift)
class ShiftAdmin(admin.ModelAdmin):
    """Admin for staff shifts."""

    list_display = ['name', 'start_time', 'end_time', 'duration_display', 'min_staff', 'max_staff', 'is_active']
    list_filter = ['is_active']
    search_fields = ['name']

    def duration_display(self, obj):
        """Show shift duration."""
        return f"{obj.duration_hours():.1f} hours"
    duration_display.short_description = 'Duration'


@admin.register(StaffShiftAssignment)
class StaffShiftAssignmentAdmin(admin.ModelAdmin):
    """Admin for staff shift assignments."""

    list_display = ['staff', 'shift', 'date', 'status', 'check_in_time', 'check_out_time']
    list_filter = ['status', 'date']
    search_fields = ['staff__first_name', 'staff__last_name', 'shift__name']
    date_hierarchy = 'date'

    actions = ['mark_confirmed', 'mark_cancelled']

    def mark_confirmed(self, request, queryset):
        """Mark assignments as confirmed."""
        count = queryset.update(status='confirmed')
        self.message_user(request, f"{count} assignments confirmed.")
    mark_confirmed.short_description = "Mark as confirmed"

    def mark_cancelled(self, request, queryset):
        """Mark assignments as cancelled."""
        count = queryset.update(status='cancelled')
        self.message_user(request, f"{count} assignments cancelled.")
    mark_cancelled.short_description = "Mark as cancelled"


@admin.register(Baggage)
class BaggageAdmin(admin.ModelAdmin):
    """Admin for baggage tracking."""

    list_display = ['tag_number', 'passenger', 'flight', 'status', 'origin', 'destination', 'weight', 'checked_in_at']
    list_filter = ['status', 'origin', 'destination', 'checked_in_at']
    search_fields = ['tag_number', 'passenger__first_name', 'passenger__last_name']
    date_hierarchy = 'checked_in_at'
    readonly_fields = ['checked_in_at', 'screened_at', 'loaded_at', 'unloaded_at', 'claimed_at']

    actions = ['mark_screened', 'mark_loaded', 'mark_unloaded', 'mark_claimed', 'mark_lost', 'bulk_delete_baggage']

    def mark_screened(self, request, queryset):
        """Mark baggage as screened."""
        from django.utils import timezone
        count = queryset.update(status='screened', screened_at=timezone.now())
        self.message_user(request, f"{count} baggage items marked as screened.")
    mark_screened.short_description = "Mark as screened"

    def mark_loaded(self, request, queryset):
        """Mark baggage as loaded."""
        from django.utils import timezone
        count = queryset.update(status='loaded', loaded_at=timezone.now())
        self.message_user(request, f"{count} baggage items marked as loaded.")
    mark_loaded.short_description = "Mark as loaded"

    def mark_unloaded(self, request, queryset):
        """Mark baggage as unloaded."""
        from django.utils import timezone
        count = queryset.update(status='unloaded', unloaded_at=timezone.now())
        self.message_user(request, f"{count} baggage items marked as unloaded.")
    mark_unloaded.short_description = "Mark as unloaded"

    def mark_claimed(self, request, queryset):
        """Mark baggage as claimed."""
        from django.utils import timezone
        count = queryset.update(status='claimed', claimed_at=timezone.now())
        self.message_user(request, f"{count} baggage items marked as claimed.")
    mark_claimed.short_description = "Mark as claimed"

    def mark_lost(self, request, queryset):
        """Mark baggage as lost."""
        count = queryset.update(status='lost')
        self.message_user(request, f"{count} baggage items marked as lost.")
    mark_lost.short_description = "Mark as lost"

    def bulk_delete_baggage(self, request, queryset):
        """Delete selected baggage records."""
        count = queryset.count()
        queryset.delete()
        self.message_user(request, f"{count} baggage records deleted successfully.")
    bulk_delete_baggage.short_description = "Delete selected records"


@admin.register(WeatherCondition)
class WeatherConditionAdmin(admin.ModelAdmin):
    """Admin for weather conditions."""

    list_display = ['airport', 'weather_description', 'severity', 'temperature', 'wind_speed', 'delay_impact', 'timestamp']
    list_filter = ['airport', 'severity', 'timestamp']
    search_fields = ['airport__code', 'weather_description']
    readonly_fields = ['fetched_at', 'created_at', 'updated_at']


@admin.register(Flight)
class FlightAdmin(admin.ModelAdmin):
    """Admin for flight management."""

    list_display = ['flight_number', 'airline', 'origin', 'destination', 'scheduled_departure', 'gate', 'status', 'delay_minutes']
    list_filter = ['status', 'origin', 'destination', 'airline']
    search_fields = ['flight_number', 'airline', 'origin', 'destination']
    date_hierarchy = 'scheduled_departure'
    
    actions = ['mark_delayed', 'mark_cancelled', 'mark_departed', 'mark_arrived', 'bulk_delete_flights']

    def mark_delayed(self, request, queryset):
        """Mark selected flights as delayed."""
        count = queryset.update(status='delayed')
        self.message_user(request, f"{count} flights marked as delayed.")
    mark_delayed.short_description = "Mark as delayed"

    def mark_cancelled(self, request, queryset):
        """Mark selected flights as cancelled."""
        count = queryset.update(status='cancelled')
        self.message_user(request, f"{count} flights marked as cancelled.")
    mark_cancelled.short_description = "Mark as cancelled"

    def mark_departed(self, request, queryset):
        """Mark selected flights as departed."""
        count = queryset.update(status='departed')
        self.message_user(request, f"{count} flights marked as departed.")
    mark_departed.short_description = "Mark as departed"

    def mark_arrived(self, request, queryset):
        """Mark selected flights as arrived."""
        count = queryset.update(status='arrived')
        self.message_user(request, f"{count} flights marked as arrived.")
    mark_arrived.short_description = "Mark as arrived"

    def bulk_delete_flights(self, request, queryset):
        """Delete selected flights."""
        count = queryset.count()
        queryset.delete()
        self.message_user(request, f"{count} flights deleted successfully.")
    bulk_delete_flights.short_description = "Delete selected flights"


@admin.register(WeatherAlert)
class WeatherAlertAdmin(admin.ModelAdmin):
    """Admin for weather alerts."""

    list_display = ['airport', 'alert_type', 'severity', 'title', 'is_active', 'start_time', 'acknowledged_at']
    list_filter = ['is_active', 'alert_type', 'severity']
    search_fields = ['airport__code', 'title', 'description']
    
    actions = ['acknowledge_alert', 'deactivate_alert']

    def acknowledge_alert(self, request, queryset):
        """Acknowledge selected alerts."""
        from django.utils import timezone
        count = queryset.update(
            acknowledged_by=request.user if request.user.is_authenticated else None,
            acknowledged_at=timezone.now()
        )
        self.message_user(request, f"{count} alerts acknowledged.")
    acknowledge_alert.short_description = "Acknowledge alerts"

    def deactivate_alert(self, request, queryset):
        """Deactivate selected alerts."""
        count = queryset.update(is_active=False)
        self.message_user(request, f"{count} alerts deactivated.")
    deactivate_alert.short_description = "Deactivate alerts"


@admin.register(FuelInventory)
class FuelInventoryAdmin(admin.ModelAdmin):
    """Admin for fuel inventory."""

    list_display = ['airport', 'fuel_type', 'storage_tank', 'current_level_display', 'capacity', 'fill_percentage_display', 'is_low']
    list_filter = ['airport', 'fuel_type']
    search_fields = ['airport__code', 'storage_tank']

    def current_level_display(self, obj):
        return f"{obj.current_level:,.0f} L"
    current_level_display.short_description = 'Current Level'

    def fill_percentage_display(self, obj):
        return f"{obj.fill_percentage:.1f}%"
    fill_percentage_display.short_description = 'Fill %'

    def is_low(self, obj):
        if obj.is_low:
            return True
        return False
    is_low.boolean = True
    is_low.short_description = 'Low Level'


@admin.register(FuelDispensing)
class FuelDispensingAdmin(admin.ModelAdmin):
    """Admin for fuel dispensing records."""

    list_display = ['flight', 'inventory', 'amount', 'start_time', 'end_time', 'operator', 'truck_id']
    list_filter = ['inventory', 'operator']
    search_fields = ['flight__flight_number', 'truck_id']
    readonly_fields = ['created_at', 'updated_at']


@admin.register(MaintenanceSchedule)
class MaintenanceScheduleAdmin(admin.ModelAdmin):
    """Admin for maintenance schedules."""

    list_display = ['title', 'equipment_type', 'equipment_id', 'airport', 'frequency', 'next_due', 'priority', 'status', 'assigned_to']
    list_filter = ['status', 'priority', 'frequency', 'equipment_type']
    search_fields = ['title', 'equipment_id', 'description']
    
    actions = ['mark_completed', 'mark_in_progress', 'mark_overdue']

    def mark_completed(self, request, queryset):
        """Mark schedules as completed."""
        count = queryset.update(status='completed')
        self.message_user(request, f"{count} schedules marked as completed.")
    mark_completed.short_description = "Mark as completed"

    def mark_in_progress(self, request, queryset):
        """Mark schedules as in progress."""
        count = queryset.update(status='in_progress')
        self.message_user(request, f"{count} schedules marked as in progress.")
    mark_in_progress.short_description = "Mark as in progress"

    def mark_overdue(self, request, queryset):
        """Mark schedules as overdue."""
        count = queryset.update(status='overdue')
        self.message_user(request, f"{count} schedules marked as overdue.")
    mark_overdue.short_description = "Mark as overdue"


@admin.register(MaintenanceLog)
class MaintenanceLogAdmin(admin.ModelAdmin):
    """Admin for maintenance logs."""

    list_display = ['equipment_id', 'equipment_type', 'maintenance_type', 'status', 'started_at', 'completed_at', 'performed_by']
    list_filter = ['equipment_type', 'maintenance_type', 'status', 'performed_by']
    search_fields = ['equipment_id', 'description', 'notes']
    readonly_fields = ['created_at', 'updated_at']


@admin.register(CustomField)
class CustomFieldAdmin(admin.ModelAdmin):
    """Admin for custom field definitions."""

    list_display = ['name', 'label', 'model_name', 'field_type', 'required', 'is_active', 'order']
    list_filter = ['model_name', 'field_type', 'is_active', 'required']
    search_fields = ['name', 'label', 'model_name']
    ordering = ['model_name', 'order', 'name']
    
    fieldsets = (
        ('Basic Information', {
            'fields': ('name', 'label', 'model_name', 'field_type')
        }),
        ('Configuration', {
            'fields': ('required', 'default_value', 'help_text', 'is_active', 'order')
        }),
        ('Choices', {
            'fields': ('choices',),
            'description': 'For choice fields, enter a JSON list like: ["Option 1", "Option 2"]'
        }),
    )


@admin.register(CustomFieldValue)
class CustomFieldValueAdmin(admin.ModelAdmin):
    """Admin for custom field values."""

    list_display = ['field', 'model_name', 'object_id', 'value']
    list_filter = ['model_name', 'field']
    search_fields = ['object_id', 'value']
    readonly_fields = ['created_at', 'updated_at']
